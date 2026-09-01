"""操作日志 API。

- GET /logs            分页查询,支持用户名/模块/操作类型/时间范围筛选
- GET /logs/modules    已出现模块名列表(筛选下拉用)
- GET /logs/export     按当前筛选条件导出 Excel
"""
import io as _io
from datetime import date, datetime, time as dt_time
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, Query, Response
from sqlmodel import Session, func, select

from app.core.database import get_session
from app.core.permissions import require_admin
from app.models import SysOperationLog, SysUser

router = APIRouter()


def _build_filter_stmt(
    session: Session,
    username: str | None,
    module: str | None,
    operation: str | None,
    start_date: date | None,
    end_date: date | None,
):
    """按筛选条件构造日志查询语句(不含排序/分页)。"""
    stmt = select(SysOperationLog)

    if username and username.strip():
        user_ids = session.exec(
            select(SysUser.user_id).where(SysUser.username.contains(username.strip()))
        ).all()
        # 无匹配用户时用不存在的 id 保证结果为空
        stmt = stmt.where(SysOperationLog.user_id.in_(user_ids or [-1]))
    if module:
        stmt = stmt.where(SysOperationLog.module == module)
    if operation:
        stmt = stmt.where(SysOperationLog.operation == operation)
    if start_date:
        stmt = stmt.where(
            SysOperationLog.operation_time >= datetime.combine(start_date, dt_time.min)
        )
    if end_date:
        stmt = stmt.where(
            SysOperationLog.operation_time <= datetime.combine(end_date, dt_time.max)
        )
    return stmt


def _serialize_log(session: Session, log: SysOperationLog) -> dict:
    user = session.get(SysUser, log.user_id)
    return {
        "id": log.log_id,
        "username": user.username if user else "",
        "operation": log.content,
        "type": log.module,
        "operationType": log.operation,
        "ip": log.ip_address or "",
        "time": log.operation_time.strftime("%Y-%m-%d %H:%M:%S") if log.operation_time else "",
    }


@router.get("/logs/modules", tags=["系统日志"])
def list_log_modules(
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(require_admin),
) -> dict:
    """返回已出现过的操作模块名列表,供筛选下拉框使用。"""
    modules = session.exec(
        select(SysOperationLog.module).distinct().order_by(SysOperationLog.module)
    ).all()
    return {"list": modules}


@router.get("/logs", tags=["系统日志"])
def list_logs(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    username: str | None = Query(default=None, description="操作用户名模糊搜索"),
    module: str | None = Query(default=None, description="操作模块精确筛选"),
    operation: str | None = Query(default=None, description="操作类型精确筛选"),
    start_date: date | None = Query(default=None, description="开始日期(含)"),
    end_date: date | None = Query(default=None, description="结束日期(含)"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(require_admin),
) -> dict:
    """分页查询操作日志,支持用户名/模块/操作类型/时间范围筛选。"""
    base = _build_filter_stmt(session, username, module, operation, start_date, end_date)
    total = session.exec(select(func.count()).select_from(base.subquery())).one()

    logs = session.exec(
        base.order_by(
            SysOperationLog.operation_time.desc(), SysOperationLog.log_id.desc()
        )
        .offset((page - 1) * page_size)
        .limit(page_size)
    ).all()

    return {"list": [_serialize_log(session, log) for log in logs], "total": total}


@router.get("/logs/export", tags=["系统日志"])
def export_logs(
    username: str | None = Query(default=None, description="操作用户名模糊搜索"),
    module: str | None = Query(default=None, description="操作模块精确筛选"),
    operation: str | None = Query(default=None, description="操作类型精确筛选"),
    start_date: date | None = Query(default=None, description="开始日期(含)"),
    end_date: date | None = Query(default=None, description="结束日期(含)"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(require_admin),
) -> Response:
    """将筛选后的全部操作日志导出为 Excel 文件(不含分页限制)。"""
    base = _build_filter_stmt(session, username, module, operation, start_date, end_date)
    logs = session.exec(
        base.order_by(
            SysOperationLog.operation_time.desc(), SysOperationLog.log_id.desc()
        )
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "操作日志"

    header_font = openpyxl.styles.Font(bold=True, size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    headers = ["序号", "操作用户", "操作模块", "操作类型", "操作内容", "IP 地址", "操作时间"]
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for ri, log in enumerate(logs, start=2):
        user = session.get(SysUser, log.user_id)
        vals = [
            ri - 1,
            user.username if user else "",
            log.module,
            log.operation,
            log.content,
            log.ip_address or "",
            log.operation_time.strftime("%Y-%m-%d %H:%M:%S") if log.operation_time else "",
        ]
        for ci, v in enumerate(vals, start=1):
            ws.cell(row=ri, column=ci, value=v)

    # 调整列宽(按列语义:序号/用户/模块/类型/内容/IP/时间)
    for ci, width in zip(range(1, 8), [8, 16, 14, 12, 50, 16, 22]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = "操作日志导出.xlsx"
    encoded = quote(safe_name, safe="")
    ascii_name = "operation_logs_export.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{ascii_name}\"; "
                f"filename*=UTF-8''{encoded}"
            ),
        },
    )
