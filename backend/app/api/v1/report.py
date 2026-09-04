"""报告生成 API（Report.Generate / Report.Export / Report.Preview）。

接口：
    GET  /api/v1/report           生成报告 JSON
    POST /api/v1/report/history   生成并保存报告快照
    GET  /api/v1/report/history   查询当前用户的报告历史
    GET  /api/v1/report/history/{id} 读取历史快照
    GET  /api/v1/report/history/{id}/download 下载历史快照
    GET  /api/v1/report/preview   在线预览 HTML
    GET  /api/v1/report/export    导出 Excel（format=xlsx）

数据源仅含成绩、考勤、课堂互动及教师发布题目的答题数据。
"""

from __future__ import annotations

import io as _io
import json
from datetime import datetime
from html import escape
from urllib.parse import quote

import httpx
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.params import Query as QueryParam
from fastapi.responses import HTMLResponse, Response
from loguru import logger
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlmodel import Session, select

from app.core.database import get_session
from app.core.operation_log import get_current_user
from app.models import ClassInfo, Course, ReportHistory, Student, SysRole, SysUser, SysOperationLog
from app.api.v1.analysis import _check_course_access
from app.services.report_template import (
    build_class_context,
    build_student_context,
    render_report,
)

router = APIRouter()

ALGO_BASE = "http://127.0.0.1:8001"

_REPORT_TYPE_SCOPE: dict[int, str] = {
    1: "class",
    2: "student",
    3: "class",
    4: "class",
}

_REPORT_TYPE_NAMES: dict[int, str] = {
    1: "班级学情分析报告",
    2: "学生个人学情报告",
    3: "课程知识点分析报告",
    4: "学生学习质量报告",
}


def _report_history_payload(
    *,
    report_type: int,
    course_id: int,
    course_name: str,
    class_id: int | None,
    student_id: int | None,
    ctx,
) -> str:
    target_name = ctx.student_name or ctx.class_name or ""
    type_name = _REPORT_TYPE_NAMES.get(report_type, "报告")
    name_parts = [part for part in [target_name, course_name, type_name] if part]
    payload = {
        "name": " - ".join(name_parts)[:90],
        "type": type_name[:12],
        "reportType": report_type,
        "courseId": course_id,
        "classId": class_id,
        "studentId": student_id,
        "format": "PDF/Excel",
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))

# 学生可访问的报告类型（仅限查看本人数据）
_STUDENT_REPORT_TYPES = {2, 3, 4}


class GenerateSavedReportRequest(BaseModel):
    course_id: int
    report_type: int = Field(ge=1, le=4)
    class_id: int | None = None
    student_id: int | None = None
    semester: str | None = Field(default=None, max_length=32)
    export_format: str = Field(default="pdf", pattern="^(pdf|xlsx)$")
    use_llm: bool = True
    dashboard_stats: dict = Field(default_factory=dict)


# ============================================================================
# 权限校验
# ============================================================================

def _check_report_access(
    session: Session,
    current_user: SysUser,
    course_id: int,
    report_type: int,
    student_id: int | None,
) -> None:
    """报告生成权限校验（Report.UserValid）。

    - 任课教师（teacher）：仅可查看自己授课课程的报告
    - 学生（student）：仅可生成类型 2（个人学情）/ 4（学习质量），且只能看自己
    - 管理员（admin）：不接触教学报告
    """
    role = session.get(SysRole, current_user.role_id)
    role_code = role.role_code if role else ""

    if role_code == "teacher":
        _check_course_access(session, current_user, course_id)
        return

    if role_code == "student":
        if report_type not in _STUDENT_REPORT_TYPES:
            raise HTTPException(
                status_code=403,
                detail="学生仅可生成个人学情报告（类型2）和学习质量报告（类型4）",
            )
        student = session.exec(
            select(Student).where(Student.user_id == current_user.user_id)
        ).first()
        if not student:
            raise HTTPException(status_code=403, detail="当前账号未关联学生信息")
        if student_id is not None and student_id != student.student_id:
            raise HTTPException(status_code=403, detail="学生仅可生成自己的报告")
        # 校验课程：学生必须选修了该课程
        from app.models import CourseStudent
        enrollment = session.exec(
            select(CourseStudent).where(
                CourseStudent.student_id == student.student_id,
                CourseStudent.course_id == course_id,
            )
        ).first()
        if not enrollment:
            raise HTTPException(
                status_code=403,
                detail="您未选修该课程，无法生成报告",
            )
        return

    raise HTTPException(status_code=403, detail="无权生成报告")


# ============================================================================
# 工具函数
# ============================================================================

def _enhance_with_llm(scope: str, report_type: int, ctx_dict: dict, template: dict) -> dict:
    """调 algorithm 服务做 LLM 增强，失败回退模板。"""
    type_name = _REPORT_TYPE_NAMES.get(report_type, "报告")
    try:
        resp = httpx.post(
            f"{ALGO_BASE}/generate_report",
            json={
                "scope": scope,
                "report_type": report_type,
                "report_type_name": type_name,
                "template": template,
                "context": ctx_dict,
            },
            timeout=30.0,
        )
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        logger.warning(f"LLM report enhancement failed, using template: {e}")
        return {**template, "source": "template_fallback", "error": str(e)[:200]}


def _ctx_to_dict(ctx) -> dict:
    """ReportContext → dict。"""
    return {
        "scope": ctx.scope,
        "course_id": ctx.course_id,
        "course_name": ctx.course_name,
        "student_id": ctx.student_id,
        "student_name": ctx.student_name,
        "class_name": ctx.class_name,
        "avg_score": ctx.avg_score,
        "pass_rate": ctx.pass_rate,
        "excellent_rate": ctx.excellent_rate,
        "weak_points": ctx.weak_points,
        "strong_points": ctx.strong_points,
        "trend": ctx.trend,
        "risk_count": ctx.risk_count,
        "tags": ctx.tags,
        "radar": ctx.radar,
        "knowledge_mastery": ctx.knowledge_mastery,
        "score_buckets": ctx.score_buckets,
        "student_count": ctx.student_count,
        "attendance_rate": ctx.attendance_rate,
        "latest_exam_name": ctx.latest_exam_name,
        "score_min": ctx.score_min,
        "score_max": ctx.score_max,
        "score_median": ctx.score_median,
        "warnings": ctx.warnings,
        "evaluation": ctx.evaluation,
        "score_history": ctx.score_history,
        "predicted_score": ctx.predicted_score,
        "findings": ctx.findings,
        "eval_snapshot": ctx.eval_snapshot,
        "exam_batches": ctx.exam_batches,
        "report_type": ctx.report_type,
    }


def _charts_from_ctx(ctx) -> dict:
    """结构化图表数据；focus 控制前端按报告主题展示不同图。"""
    snapshot = ctx.eval_snapshot or {}
    indexes = []
    for dim in snapshot.get("scheme") or []:
        for item in dim.get("indexes") or []:
            if item.get("score") is None:
                continue
            indexes.append({
                "name": f"{dim.get('name', '')}-{item.get('name', '')}",
                "score": item["score"],
                "weight": item.get("weight", 0),
            })
    focus = {1: "class", 2: "student", 3: "knowledge", 4: "quality"}.get(ctx.report_type, "class")
    return {
        "focus": focus,
        "scoreBuckets": ctx.score_buckets or [],
        "knowledge": ctx.knowledge_mastery or [],
        "radar": ctx.radar or {},
        "evalIndexes": indexes,
        "academicParts": snapshot.get("academicParts") or [],
        "rates": {
            "avgScore": ctx.avg_score,
            "passRate": ctx.pass_rate,
            "excellentRate": ctx.excellent_rate,
            "attendanceRate": ctx.attendance_rate,
        },
        "scoreHistory": ctx.score_history or [],
    }


def _merge_saved_stats(dashboard_stats: dict | None, report: dict) -> dict:
    """看板快照与本次报告指标合并：空值不覆盖已有数据。"""
    merged = dict(dashboard_stats or {})
    for key, value in (report.get("metrics") or {}).items():
        if value in (None, ""):
            continue
        existing = merged.get(key)
        if value in (0, 0.0) and existing not in (None, "", 0, 0.0):
            continue
        merged[key] = value
    merged["charts"] = report.get("charts") or {}
    return merged


def _assemble_report(
    session: Session,
    course_id: int,
    report_type: int,
    class_id: int | None,
    student_id: int | None,
    use_llm: bool,
    current_user: SysUser | None = None,
) -> tuple[dict, any]:
    """组装报告 JSON 和上下文对象。返回 (report_dict, ReportContext)。"""
    scope = _REPORT_TYPE_SCOPE.get(report_type, "class")
    type_name = _REPORT_TYPE_NAMES.get(report_type, "报告")

    # Unwrap Query params（直接 Python 调用兼容）
    _class_id = class_id if not isinstance(class_id, QueryParam) else None
    _student_id = student_id if not isinstance(student_id, QueryParam) else None

    # 学生角色：类型 2/4 自动取本人 student_id
    if current_user:
        role = session.get(SysRole, current_user.role_id)
        role_code = role.role_code if role else ""
        if role_code == "student" and report_type in _STUDENT_REPORT_TYPES:
            student = session.exec(
                select(Student).where(Student.user_id == current_user.user_id)
            ).first()
            if student:
                _student_id = student.student_id
                scope = "student"

    if report_type == 2:
        if not _student_id:
            raise HTTPException(status_code=400, detail="学生个人报告必须提供 student_id")
        ctx = build_student_context(session, _student_id, course_id, report_type)
    elif report_type in (3, 4) and _student_id:
        ctx = build_student_context(session, _student_id, course_id, report_type)
        scope = "student"
    else:
        ctx = build_class_context(session, course_id, _class_id, report_type)

    ctx.scope = scope
    template = render_report(ctx)

    charts = _charts_from_ctx(ctx)
    detail = {
        "findings": template.get("findings") or ctx.findings,
        "metrics": template.get("metrics") or {},
        "warnings": template.get("warnings") if "warnings" in template else ctx.warnings,
        "evalScheme": template.get("evalScheme") or [],
        "academicParts": template.get("academicParts") or [],
        "charts": charts,
        "report_type": report_type,
        "report_type_name": type_name,
    }
    if not use_llm:
        return {**template, **detail}, ctx

    ctx_dict = _ctx_to_dict(ctx)
    ctx_dict["report_type"] = report_type
    ctx_dict["report_type_name"] = type_name
    enhanced = _enhance_with_llm(scope, report_type, ctx_dict, template)
    return {**enhanced, **detail}, ctx


def _history_to_dict(history: ReportHistory, include_snapshot: bool = False) -> dict:
    """Return the stable API representation of a persisted report."""
    result = {
        "id": history.report_id,
        "name": history.report_name,
        "type": _REPORT_TYPE_NAMES.get(history.report_type, "报告"),
        "report_type": history.report_type,
        "scope": history.scope,
        "format": history.export_format.upper(),
        "time": history.created_at.strftime("%Y-%m-%d %H:%M"),
        "created_at": history.created_at.isoformat(),
        "course_id": history.course_id,
        "course_name": history.course_name,
        "class_id": history.class_id,
        "class_name": history.class_name,
        "student_id": history.student_id,
        "student_name": history.student_name,
    }
    if include_snapshot:
        result["parameters"] = json.loads(history.parameter_snapshot)
        result["data"] = json.loads(history.report_snapshot)
        result["stats"] = json.loads(history.stats_snapshot)
    return result


def _get_owned_history(
    session: Session,
    report_id: int,
    current_user: SysUser,
) -> ReportHistory:
    history = session.get(ReportHistory, report_id)
    if not history:
        raise HTTPException(status_code=404, detail="报告不存在")
    if history.creator_user_id != current_user.user_id:
        raise HTTPException(status_code=403, detail="无权访问该报告")
    _check_report_access(
        session,
        current_user,
        history.course_id,
        history.report_type,
        history.student_id,
    )
    return history


def _snapshot_workbook(history: ReportHistory) -> bytes:
    """Build a readable xlsx from the immutable history snapshot."""
    report = json.loads(history.report_snapshot)
    stats = json.loads(history.stats_snapshot)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报告正文"

    title_fill = openpyxl.styles.PatternFill(
        start_color="2563EB", end_color="2563EB", fill_type="solid"
    )
    title_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=12)
    wrap = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A1:B1")
    ws["A1"] = history.report_name
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=16)
    ws["A2"] = "课程"
    ws["B2"] = history.course_name
    ws["A3"] = "生成时间"
    ws["B3"] = history.created_at.strftime("%Y-%m-%d %H:%M")
    rows = [
        ("总体概述", report.get("summary", "")),
        ("关键结论", report.get("conclusion", "")),
        ("建议措施", report.get("suggestion", "")),
    ]
    for row_index, (label, value) in enumerate(rows, start=5):
        ws.cell(row=row_index, column=1, value=label)
        ws.cell(row=row_index, column=2, value=value)
        ws.cell(row=row_index, column=1).font = title_font
        ws.cell(row=row_index, column=1).fill = title_fill
        ws.cell(row=row_index, column=2).alignment = wrap
        ws.row_dimensions[row_index].height = 132
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 90

    if stats:
        stats_ws = wb.create_sheet("指标快照")
        stats_ws.append(["指标", "数值"])
        for cell in stats_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        skip_keys = {"charts"}
        for key, value in stats.items():
            if key in skip_keys or isinstance(value, (dict, list)):
                continue
            stats_ws.append([key, value])
        stats_ws.column_dimensions["A"].width = 24
        stats_ws.column_dimensions["B"].width = 24

    charts = report.get("charts") or stats.get("charts") or {}
    score_buckets = [item for item in (charts.get("scoreBuckets") or []) if item.get("count")]
    knowledge = charts.get("knowledge") or []
    radar = charts.get("radar") or {}

    if score_buckets:
        dist_ws = wb.create_sheet("成绩分布")
        dist_ws.append(["成绩段", "人数"])
        for cell in dist_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        for item in score_buckets:
            dist_ws.append([item.get("label", ""), item.get("count", 0)])
        dist_ws.column_dimensions["A"].width = 22
        dist_ws.column_dimensions["B"].width = 12
        pie = PieChart()
        pie.title = "成绩分布"
        labels = Reference(dist_ws, min_col=1, min_row=2, max_row=1 + len(score_buckets))
        data = Reference(dist_ws, min_col=2, min_row=1, max_row=1 + len(score_buckets))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showCatName = True
        pie.width = 14
        pie.height = 8
        dist_ws.add_chart(pie, "D2")

    if knowledge:
        kp_ws = wb.create_sheet("知识点掌握")
        kp_ws.append(["知识点", "掌握度(%)", "等级"])
        for cell in kp_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        for item in knowledge:
            kp_ws.append([item.get("name", ""), item.get("accuracy", 0), item.get("level", "")])
        kp_ws.column_dimensions["A"].width = 24
        kp_ws.column_dimensions["B"].width = 14
        kp_ws.column_dimensions["C"].width = 12
        bar = BarChart()
        bar.type = "bar"
        bar.title = "知识点掌握度"
        bar.y_axis.title = None
        bar.x_axis.title = "掌握度(%)"
        bar.x_axis.scaling.max = 100
        data = Reference(kp_ws, min_col=2, min_row=1, max_row=1 + len(knowledge))
        cats = Reference(kp_ws, min_col=1, min_row=2, max_row=1 + len(knowledge))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.shape = 4
        bar.legend = None
        bar.width = 16
        bar.height = max(8, min(18, 1.2 * len(knowledge)))
        kp_ws.add_chart(bar, "E2")

    if radar:
        radar_ws = wb.create_sheet("能力雷达")
        radar_ws.append(["维度", "得分"])
        for cell in radar_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        for name, value in radar.items():
            radar_ws.append([name, value])
        radar_ws.column_dimensions["A"].width = 18
        radar_ws.column_dimensions["B"].width = 12
        radar_bar = BarChart()
        radar_bar.type = "col"
        radar_bar.title = "能力维度得分"
        radar_bar.y_axis.scaling.max = 100
        data = Reference(radar_ws, min_col=2, min_row=1, max_row=1 + len(radar))
        cats = Reference(radar_ws, min_col=1, min_row=2, max_row=1 + len(radar))
        radar_bar.add_data(data, titles_from_data=True)
        radar_bar.set_categories(cats)
        radar_bar.legend = None
        radar_bar.width = 12
        radar_bar.height = 8
        radar_ws.add_chart(radar_bar, "D2")

    scheme_rows = report.get("evalScheme") or []
    if scheme_rows:
        scheme_ws = wb.create_sheet("教师评价方案")
        scheme_ws.append(["维度", "维度得分", "指标", "权重(%)", "指标得分"])
        for cell in scheme_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        for dim in scheme_rows:
            indexes = dim.get("indexes") or [{}]
            for item in indexes:
                scheme_ws.append([
                    dim.get("name", ""),
                    dim.get("score"),
                    item.get("name", ""),
                    item.get("weight"),
                    item.get("score"),
                ])
        scheme_ws.column_dimensions["A"].width = 16
        scheme_ws.column_dimensions["B"].width = 12
        scheme_ws.column_dimensions["C"].width = 18
        scheme_ws.column_dimensions["D"].width = 12
        scheme_ws.column_dimensions["E"].width = 12

    findings = report.get("findings") or []
    if findings:
        find_ws = wb.create_sheet("现状要点")
        find_ws.append(["序号", "要点"])
        for cell in find_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        for index, text in enumerate(findings, start=1):
            find_ws.append([index, text])
            find_ws.cell(row=index + 1, column=2).alignment = wrap
            find_ws.row_dimensions[index + 1].height = 36
        find_ws.column_dimensions["A"].width = 8
        find_ws.column_dimensions["B"].width = 100

    warning_rows = report.get("warnings") or []
    if warning_rows:
        warn_ws = wb.create_sheet("预警名单")
        warn_ws.append(["学生", "等级", "原因"])
        for cell in warn_ws[1]:
            cell.font = title_font
            cell.fill = title_fill
        for item in warning_rows:
            warn_ws.append([
                item.get("name", ""),
                item.get("level", ""),
                "；".join(item.get("reasons") or []),
            ])
        warn_ws.column_dimensions["A"].width = 16
        warn_ws.column_dimensions["B"].width = 10
        warn_ws.column_dimensions["C"].width = 70

    output = _io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _snapshot_pdf(history: ReportHistory) -> bytes:
    """Build a paginated, Chinese-readable PDF from an immutable snapshot."""
    report = json.loads(history.report_snapshot)
    stats = json.loads(history.stats_snapshot)
    output = _io.BytesIO()

    font_name = "STSong-Light"
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont(font_name))

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontName=font_name,
        fontSize=20, leading=28, alignment=TA_CENTER, textColor=colors.HexColor("#0f172a"),
        spaceAfter=8 * mm,
    )
    meta_style = ParagraphStyle(
        "ReportMeta", parent=styles["Normal"], fontName=font_name,
        fontSize=10, leading=16, alignment=TA_CENTER, textColor=colors.HexColor("#64748b"),
        spaceAfter=8 * mm,
    )
    heading_style = ParagraphStyle(
        "ReportHeading", parent=styles["Heading2"], fontName=font_name,
        fontSize=14, leading=20, textColor=colors.HexColor("#1e40af"),
        spaceBefore=5 * mm, spaceAfter=3 * mm, keepWithNext=True,
    )
    body_style = ParagraphStyle(
        "ReportBody", parent=styles["BodyText"], fontName=font_name,
        fontSize=11, leading=20, textColor=colors.HexColor("#1f2937"),
        wordWrap="CJK", spaceAfter=3 * mm,
    )

    def _paragraph(value: object, style: ParagraphStyle) -> Paragraph:
        text = escape(str(value or "-")).replace("\n", "<br/>")
        return Paragraph(text, style)

    story = [
        _paragraph(history.report_name, title_style),
        _paragraph(
            f"课程：{history.course_name}　生成时间：{history.created_at:%Y-%m-%d %H:%M}",
            meta_style,
        ),
    ]

    stat_labels = {
        "studentCount": "学生人数", "avgScore": "均分",
        "scoreMedian": "中位数", "scoreMin": "最低分", "scoreMax": "最高分",
        "passRate": "及格率", "excellentRate": "优秀率",
        "attendanceRate": "平均出勤率", "warningCount": "预警学生",
        "latestExam": "最近考核", "courseCount": "课程数量",
    }
    has_core_stats = bool(stats) and any(key in stats for key in stat_labels)
    charts = report.get("charts") or (stats.get("charts") if isinstance(stats, dict) else {}) or {}
    focus = charts.get("focus") or {1: "class", 2: "student", 3: "knowledge", 4: "quality"}.get(history.report_type, "class")
    score_buckets = [item for item in (charts.get("scoreBuckets") or []) if item.get("count")] if focus == "class" else []
    knowledge = charts.get("knowledge") or [] if focus in ("knowledge", "student") else []
    radar = charts.get("radar") or {} if focus in ("student", "quality") else {}
    has_charts = bool(score_buckets or knowledge or radar or (charts.get("evalIndexes") if focus == "quality" else None))
    if focus == "knowledge":
        has_core_stats = False

    def _styled_table(table_data: list[list[str]]) -> Table:
        table = Table(table_data, colWidths=[55 * mm, 90 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 7),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        return table

    section_index = 0
    cn_nums = ("一", "二", "三", "四", "五", "六", "七", "八")

    if has_core_stats:
        story.append(_paragraph(f"{cn_nums[section_index]}、核心指标概览", heading_style))
        section_index += 1
        table_data = [["指标", "数值"]]
        for key, label in stat_labels.items():
            if key in stats:
                suffix = "%" if key in {"passRate", "excellentRate", "attendanceRate"} else ""
                table_data.append([label, f"{stats[key]}{suffix}"])
        story.extend([_styled_table(table_data), Spacer(1, 3 * mm)])

    if has_charts:
        story.append(_paragraph(f"{cn_nums[section_index]}、图形化数据", heading_style))
        section_index += 1
        if score_buckets:
            story.append(_paragraph("成绩分布", body_style))
            dist_rows = [["成绩段", "人数"]]
            dist_rows.extend([[item.get("label", ""), str(item.get("count", 0))] for item in score_buckets])
            story.extend([_styled_table(dist_rows), Spacer(1, 2 * mm)])
        if knowledge:
            story.append(_paragraph("知识点掌握度", body_style))
            kp_rows = [["知识点", "掌握度"]]
            kp_rows.extend([
                [item.get("name", ""), f"{item.get('accuracy', 0)}%"]
                for item in knowledge
            ])
            story.extend([_styled_table(kp_rows), Spacer(1, 2 * mm)])
        if radar:
            story.append(_paragraph("能力维度", body_style))
            radar_rows = [["维度", "得分"]]
            radar_rows.extend([[str(name), str(value)] for name, value in radar.items()])
            story.extend([_styled_table(radar_rows), Spacer(1, 2 * mm)])

    scheme_rows = report.get("evalScheme") or []
    if scheme_rows:
        story.append(_paragraph(f"{cn_nums[section_index]}、教师评价方案", heading_style))
        section_index += 1
        scheme_table = [["维度", "指标 / 权重 / 得分"]]
        for dim in scheme_rows:
            indexes = dim.get("indexes") or []
            index_text = "；".join(
                f"{item.get('name', '')} {item.get('weight', 0)}%"
                + (f" / {item['score']}分" if item.get("score") is not None else "")
                for item in indexes
            ) or "未设指标"
            dim_score = dim.get("score")
            dim_label = f"{dim.get('name', '')}" + (f"（{dim_score}分）" if dim_score is not None else "")
            scheme_table.append([dim_label, index_text])
        story.extend([_styled_table(scheme_table), Spacer(1, 2 * mm)])

    findings = [str(item) for item in (report.get("findings") or []) if item]
    if findings:
        story.append(_paragraph(f"{cn_nums[section_index]}、现状要点", heading_style))
        section_index += 1
        for index, text in enumerate(findings, start=1):
            story.append(_paragraph(f"{index}. {text}", body_style))

    warning_rows = report.get("warnings") or []
    if warning_rows:
        story.append(_paragraph(f"{cn_nums[section_index]}、预警学生", heading_style))
        section_index += 1
        warn_table = [["学生", "等级 / 原因"]]
        for item in warning_rows:
            warn_table.append([
                item.get("name", ""),
                f"{item.get('level', '')}：{'；'.join(item.get('reasons') or [])}",
            ])
        story.extend([_styled_table(warn_table), Spacer(1, 2 * mm)])

    for label, key in (("总体概述", "summary"), ("关键结论", "conclusion"), ("建议措施", "suggestion")):
        story.append(_paragraph(f"{cn_nums[section_index]}、{label}", heading_style))
        section_index += 1
        story.append(_paragraph(report.get(key, ""), body_style))

    def _draw_footer(canvas, doc):
        canvas.saveState()
        canvas.setFont(font_name, 9)
        canvas.setFillColor(colors.HexColor("#94a3b8"))
        canvas.drawCentredString(A4[0] / 2, 12 * mm, f"第 {doc.page} 页")
        canvas.restoreState()

    document = SimpleDocTemplate(
        output, pagesize=A4, leftMargin=22 * mm, rightMargin=22 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
        title=history.report_name, author="AI 辅助教学评价系统",
    )
    document.build(story, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
    return output.getvalue()


# ============================================================================
# 1. 报告历史 —— Report.History
# ============================================================================

# 注：报告历史列表查询接口定义在下方「报告历史持久化」章节
# （GET /report/history，读取 ReportHistory 表），此处不再重复注册，
# 避免 FastAPI 因重复路径而用旧版（查 SysOperationLog）遮蔽新版。

# ============================================================================
# 1. 报告生成（JSON）—— Report.Generate
# ============================================================================
@router.get("/report", tags=["报告生成"])
def get_report(
    course_id: int = Query(..., description="课程 ID"),
    report_type: int = Query(..., ge=1, le=4, description="报告类型：1=班级学情 2=学生个人 3=课程知识点 4=学习质量"),
    class_id: int | None = Query(default=None),
    student_id: int | None = Query(default=None),
    use_llm: bool = Query(default=True, description="是否使用 LLM 增强"),
    record_history: bool = Query(default=True, description="是否写入报告生成历史"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """统一报告生成接口（Report.Generate）。

    权限（Report.UserValid）：
    - 管理员不接触教学报告
    - 任课教师可生成本课程报告
    - 学生可生成类型 2/4 的个人报告（仅限本人）
    """
    _check_report_access(session, current_user, course_id, report_type, student_id)

    report, ctx = _assemble_report(session, course_id, report_type, class_id, student_id, use_llm, current_user)
    if record_history:
        course = session.get(Course, course_id)
        session.add(SysOperationLog(
            user_id=current_user.user_id or 0,
            module="报告生成",
            operation="生成",
            content=_report_history_payload(
                report_type=report_type,
                course_id=course_id,
                course_name=course.course_name if course else "",
                class_id=class_id if not isinstance(class_id, QueryParam) else None,
                student_id=ctx.student_id or (student_id if not isinstance(student_id, QueryParam) else None),
                ctx=ctx,
            ),
        ))
        session.commit()
    return report


@router.post("/report/history", tags=["报告生成"])
def generate_and_save_report(
    payload: GenerateSavedReportRequest,
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """Generate a report and persist an immutable content/parameter snapshot."""
    _check_report_access(
        session,
        current_user,
        payload.course_id,
        payload.report_type,
        payload.student_id,
    )
    report, ctx = _assemble_report(
        session,
        payload.course_id,
        payload.report_type,
        payload.class_id,
        payload.student_id,
        payload.use_llm,
        current_user,
    )
    report = {
        **report,
        "scope": ctx.scope,
        "report_type": payload.report_type,
        "report_type_name": _REPORT_TYPE_NAMES[payload.report_type],
    }

    course = session.get(Course, payload.course_id)
    if not course:
        raise HTTPException(status_code=404, detail="课程不存在")
    class_info = session.get(ClassInfo, payload.class_id) if payload.class_id else None
    student = session.get(Student, ctx.student_id) if ctx.student_id else None
    subject_name = (
        student.real_name if student else class_info.class_name if class_info else ctx.class_name
    )
    type_name = _REPORT_TYPE_NAMES[payload.report_type]
    report_name = " - ".join(part for part in (subject_name, course.course_name, type_name) if part)

    parameters = {
        "course_id": payload.course_id,
        "report_type": payload.report_type,
        "class_id": payload.class_id,
        "student_id": ctx.student_id,
        "semester": payload.semester,
        "use_llm": payload.use_llm,
        "export_format": payload.export_format,
    }
    history = ReportHistory(
        creator_user_id=current_user.user_id,
        course_id=payload.course_id,
        report_type=payload.report_type,
        scope=ctx.scope,
        class_id=payload.class_id,
        student_id=ctx.student_id,
        export_format=payload.export_format,
        report_name=report_name,
        course_name=course.course_name,
        class_name=class_info.class_name if class_info else ctx.class_name,
        student_name=student.real_name if student else ctx.student_name,
        parameter_snapshot=json.dumps(parameters, ensure_ascii=False),
        report_snapshot=json.dumps(report, ensure_ascii=False),
        stats_snapshot=json.dumps(_merge_saved_stats(payload.dashboard_stats, report), ensure_ascii=False),
        created_at=datetime.now(),
    )
    session.add(history)
    session.commit()
    session.refresh(history)
    return _history_to_dict(history, include_snapshot=True)


@router.get("/report/history", tags=["报告生成"])
def list_report_history(
    limit: int = Query(default=50, ge=1, le=200),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> list[dict]:
    """List reports created by the current user, newest first."""
    histories = session.exec(
        select(ReportHistory)
        .where(ReportHistory.creator_user_id == current_user.user_id)
        .order_by(ReportHistory.created_at.desc())
        .limit(limit)
    ).all()
    return [_history_to_dict(history) for history in histories]


@router.get("/report/history/{report_id}", tags=["报告生成"])
def get_report_history(
    report_id: int,
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """Read one persisted report after ownership and course checks."""
    history = _get_owned_history(session, report_id, current_user)
    return _history_to_dict(history, include_snapshot=True)


@router.get("/report/history/{report_id}/download", tags=["报告生成"])
def download_report_history(
    report_id: int,
    format: str = Query(default="xlsx", pattern="^(pdf|xlsx)$"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> Response:
    """Download a persisted snapshot as a native PDF or Excel workbook."""
    history = _get_owned_history(session, report_id, current_user)
    if format == "pdf":
        content = _snapshot_pdf(history)
        media_type = "application/pdf"
    else:
        content = _snapshot_workbook(history)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    safe_name = f"{history.report_name}.{format}"
    encoded = quote(safe_name, safe="")
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": (
                f'attachment; filename="report_{history.report_id}.{format}"; '
                f"filename*=UTF-8''{encoded}"
            )
        },
    )


# ============================================================================
# 2. 在线预览（HTML）—— Report.Preview
# ============================================================================

_PREVIEW_CSS = """
<style>
  body { font-family: 'Microsoft YaHei', sans-serif; max-width: 800px; margin: 40px auto; padding: 20px; color: #333; }
  h1 { text-align: center; font-size: 22px; border-bottom: 2px solid #1890ff; padding-bottom: 12px; }
  h2 { font-size: 16px; color: #1890ff; margin-top: 28px; border-left: 4px solid #1890ff; padding-left: 10px; }
  .meta { text-align: center; color: #888; font-size: 13px; margin-bottom: 24px; }
  .card { background: #fafafa; border-radius: 8px; padding: 16px 20px; margin: 12px 0; line-height: 1.8; }
  .stats { display: flex; gap: 16px; flex-wrap: wrap; margin: 16px 0; }
  .stat-item { flex: 1; min-width: 100px; background: #e6f7ff; border-radius: 6px; padding: 12px; text-align: center; }
  .stat-value { font-size: 24px; font-weight: bold; color: #1890ff; }
  .stat-label { font-size: 12px; color: #666; margin-top: 4px; }
  .weak { color: #f5222d; }
  .strong { color: #52c41a; }
  .footer { text-align: center; color: #bbb; font-size: 12px; margin-top: 40px; border-top: 1px solid #eee; padding-top: 16px; }
</style>
"""


@router.get("/report/preview", tags=["报告生成"])
def preview_report(
    course_id: int = Query(...),
    report_type: int = Query(..., ge=1, le=4),
    class_id: int | None = Query(default=None),
    student_id: int | None = Query(default=None),
    use_llm: bool = Query(default=False),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> HTMLResponse:
    """报告在线预览 HTML 页面（Report.Preview）。

    返回可直接在浏览器中展示的完整 HTML 页面。

    权限（Report.UserValid）：
    - 管理员不接触教学报告
    - 任课教师可预览本课程报告
    - 学生可预览类型 2/4 的个人报告（仅限本人）
    """
    _check_report_access(session, current_user, course_id, report_type, student_id)

    report, ctx = _assemble_report(session, course_id, report_type, class_id, student_id, use_llm, current_user)
    type_name = _REPORT_TYPE_NAMES.get(report_type, "报告")

    # 核心指标卡片
    stat_cards = ""
    if ctx.avg_score:
        stat_cards += f'<div class="stat-item"><div class="stat-value">{ctx.avg_score}</div><div class="stat-label">均分</div></div>'
    if ctx.pass_rate:
        stat_cards += f'<div class="stat-item"><div class="stat-value">{ctx.pass_rate}%</div><div class="stat-label">及格率</div></div>'
    if ctx.excellent_rate:
        stat_cards += f'<div class="stat-item"><div class="stat-value">{ctx.excellent_rate}%</div><div class="stat-label">优秀率</div></div>'

    weak_html = ""
    if ctx.weak_points:
        items = "".join(f'<span class="weak">{p}</span>、' for p in ctx.weak_points[:5]).rstrip("、")
        weak_html = f"<p>🔴 薄弱知识点：{items}</p>"
    strong_html = ""
    if ctx.strong_points:
        items = "".join(f'<span class="strong">{p}</span>、' for p in ctx.strong_points[:5]).rstrip("、")
        strong_html = f"<p>🟢 优势知识点：{items}</p>"

    tags_html = ""
    if ctx.tags:
        tags_html = f"<p>🏷️ 标签：{'、'.join(ctx.tags)}</p>"

    radar_html = ""
    if ctx.radar:
        rows = "".join(
            f'<tr><td>{k}</td><td style="width:200px"><progress value="{v}" max="100" style="width:100%"></progress></td><td>{v}</td></tr>'
            for k, v in ctx.radar.items()
        )
        radar_html = f'<table style="width:100%">{rows}</table>'

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>{type_name} - {ctx.course_name}</title>{_PREVIEW_CSS}</head>
<body>
<h1>{type_name}</h1>
<div class="meta">
  📚 课程：{ctx.course_name}
  {"| 👨‍🎓 学生：" + ctx.student_name if ctx.student_name else ""}
  {"| 🏫 班级：" + ctx.class_name if ctx.class_name else ""}
  | 📅 生成时间：{__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}
</div>

<h2>📊 核心指标</h2>
<div class="stats">{stat_cards}</div>

<h2>📈 分析摘要</h2>
<div class="card">{report.get("summary", "")}</div>

<h2>🔍 分析结论</h2>
<div class="card">
  {report.get("conclusion", "")}
  {weak_html}
  {strong_html}
  {tags_html}
</div>

{("<h2>🎯 能力雷达</h2><div class='card'>" + radar_html + "</div>") if radar_html else ""}

<h2>💡 优化建议</h2>
<div class="card">{report.get("suggestion", "")}</div>

<div class="footer">
  AI 辅助教学评价系统 · {type_name} · 数据源：成绩、考勤、课堂互动、答题记录
</div>
</body></html>"""

    return HTMLResponse(content=html)


# ============================================================================
# 3. Excel 导出 —— Report.Export
# ============================================================================

@router.get("/report/export", tags=["报告生成"])
def export_report(
    course_id: int = Query(...),
    report_type: int = Query(..., ge=1, le=4),
    format: str = Query(default="xlsx", description="导出格式：xlsx"),
    class_id: int | None = Query(default=None),
    student_id: int | None = Query(default=None),
    use_llm: bool = Query(default=False),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> Response:
    """报告导出为 Excel 文件（Report.Export）。

    生成结构化的 .xlsx 工作簿，包含：
    - Sheet1: 报告正文（摘要/结论/建议）
    - Sheet2: 核心指标数据
    - Sheet3: 知识点详情（如有）

    权限（Report.UserValid）：
    - 管理员不接触教学报告
    - 任课教师可导出本课程报告
    - 学生可导出类型 2/4 的个人报告（仅限本人）
    """
    _check_report_access(session, current_user, course_id, report_type, student_id)

    report, ctx = _assemble_report(session, course_id, report_type, class_id, student_id, use_llm, current_user)
    type_name = _REPORT_TYPE_NAMES.get(report_type, "报告")

    wb = openpyxl.Workbook()

    # ── 样式 ──
    header_fill = openpyxl.styles.PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    header_font_white = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
    title_font = openpyxl.styles.Font(bold=True, size=14)
    wrap_align = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    def _write_sheet(ws, rows: list[list], col_widths: list[int] | None = None):
        for ri, row in enumerate(rows, start=1):
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri == 1:
                    cell.font = header_font_white
                    cell.fill = header_fill
                cell.alignment = wrap_align
        if col_widths:
            for ci, w in enumerate(col_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # ── Sheet 1: 报告正文 ──
    ws1 = wb.active
    ws1.title = "报告正文"
    now_str = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')
    ws1.merge_cells("A1:C1")
    ws1.cell(row=1, column=1, value=type_name).font = title_font
    ws1.cell(row=2, column=1, value=f"课程：{ctx.course_name}  |  "
             f"学生：{ctx.student_name or '班级整体'}  |  "
             f"生成时间：{now_str}")

    rows1 = [
        ["章节", "标题", "内容"],
        ["一", "摘要", report.get("summary", "")],
        ["二", "结论", report.get("conclusion", "")],
        ["三", "建议", report.get("suggestion", "")],
    ]
    for ri, row in enumerate(rows1, start=3):
        for ci, val in enumerate(row, start=1):
            ws1.cell(row=ri, column=ci, value=val).alignment = wrap_align
    for ci in range(1, len(rows1[0]) + 1):
        cell = ws1.cell(row=3, column=ci)
        cell.font = header_font_white
        cell.fill = header_fill
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 10
    ws1.column_dimensions['C'].width = 80

    # ── Sheet 2: 核心指标 ──
    ws2 = wb.create_sheet("核心指标")
    rows2 = [
        ["指标", "数值", "说明"],
        ["班级均分", ctx.avg_score, "最近一次考核的班级平均分"],
        ["及格率", f"{ctx.pass_rate}%", "成绩 ≥ 60 分占比"],
        ["优秀率", f"{ctx.excellent_rate}%", "成绩 ≥ 85 分占比"],
        ["成绩趋势", ctx.trend, "基于历次成绩的回归斜率判断"],
        ["标签", "、".join(ctx.tags) if ctx.tags else "无", "自动生成的学习特征标签"],
    ]
    _write_sheet(ws2, rows2, col_widths=[16, 14, 45])

    # ── Sheet 3: 知识点详情 ──
    if ctx.weak_points or ctx.strong_points:
        ws3 = wb.create_sheet("知识点详情")
        rows3 = [["类型", "知识点"]]
        for pt in ctx.weak_points:
            rows3.append(["薄弱", pt])
        for pt in ctx.strong_points:
            rows3.append(["优势", pt])
        _write_sheet(ws3, rows3, col_widths=[10, 40])

    # ── 雷达图数据（学生报告时）──
    if ctx.radar:
        ws4 = wb.create_sheet("能力雷达")
        rows4 = [["维度", "得分(0-100)"]]
        for k, v in ctx.radar.items():
            rows4.append([k, v])
        _write_sheet(ws4, rows4, col_widths=[20, 16])

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    ctx_name = ctx.student_name or ctx.class_name or "班级"
    safe_name = f"{type_name}_{ctx.course_name}_{ctx_name}.xlsx"
    encoded = quote(safe_name, safe="")
    ascii_name = f"report_type{report_type}.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_name}"; '
                f"filename*=UTF-8''{encoded}"
            ),
        },
    )



