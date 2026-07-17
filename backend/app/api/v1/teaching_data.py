"""教学数据接口：文件上传导入、模板下载、数据查询、编辑、导出。

需求对应：
  3.2.5  模板下载    Data.Template.*
  3.2.6  文件上传    Data.FileUpload.*
  3.2.7  查询与管理  Data.Query.*
"""

import io as _io
import json
import os
import tempfile
import uuid
from datetime import datetime
from urllib.parse import quote

import openpyxl
from typing import Any
from fastapi import APIRouter, Body, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from sqlmodel import Session, select
from sqlalchemy import or_

from app.core.database import get_session
from app.core.operation_log import get_current_user
from app.models import (
    ScoreRecord, AttendanceRecord, ExamBatch, Course, Student,
    SysUser, Teacher, SysRole,
    IndividualScore, AttendanceSheet, CourseTestDetail,
)
from app.services.file_import import import_file, ImportResult, TEMPLATE_META, generate_template_xlsx, generate_template_txt
from app.services.analysis_refresh import refresh_course_analysis

router = APIRouter()

ALLOWED_EXTENSIONS = {".xlsx", ".txt"}


# ============================================================================
# 权限校验辅助
# ============================================================================

def _require_teacher_for_course(
    current_user: SysUser,
    course_id: int,
    session: Session,
) -> Teacher:
    """校验当前用户是指定课程的授课教师，返回 Teacher 记录。"""
    course = session.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="课程不存在")

    teacher = session.exec(
        select(Teacher).where(Teacher.user_id == current_user.user_id)
    ).first()
    if not teacher:
        raise HTTPException(status_code=403, detail="仅任课教师可操作，当前账号未关联教师")

    if course.teacher_id != teacher.teacher_id:
        raise HTTPException(
            status_code=403,
            detail=f"仅授课教师可操作。课程「{course.course_name}」不属于您",
        )
    return teacher


# ============================================================================
# 查询接口（3.2.7 节：Data.Query, Data.Query.Fuzzy, Data.Query.Filter）
# ============================================================================

STATUS_MAP = {0: "正常", 1: "迟到", 2: "早退", 3: "缺勤", 4: "请假"}


@router.get("/teaching-data", tags=["教学数据"])
def query_teaching_data(
    course_id: int = Query(..., description="课程 ID"),
    keyword: str | None = Query(default=None, description="学生姓名/学号模糊搜索"),
    data_type: str | None = Query(default=None, description="数据类型: score / attendance"),
    batch_id: int | None = Query(default=None, description="考核批次 ID（仅 dataType=score 时生效）"),
    page: int = Query(default=1, ge=1, description="页码"),
    page_size: int = Query(default=50, ge=1, le=200, description="每页条数"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """查询教学数据，支持模糊搜索与组合筛选。

    权限（Data.Query.UserValid）：仅授课教师可查询自己课程的数据。
    模糊搜索（Data.Query.Fuzzy）：keyword 匹配学生姓名或学号。
    组合筛选（Data.Query.Filter）：dataType + batchId。
    """
    # 权限校验
    _require_teacher_for_course(current_user, course_id, session)

    # 构建基础学生查询（模糊匹配）
    student_stmt = select(Student)
    if keyword:
        like = f"%{keyword}%"
        student_stmt = student_stmt.where(
            or_(Student.real_name.like(like), Student.student_no.like(like))
        )
    matched_students = session.exec(student_stmt).all()
    student_ids = {s.student_id for s in matched_students}

    rows: list[dict] = []

    # ── 成绩数据 ──
    if data_type is None or data_type == "score":
        # 旧表：ScoreRecord
        stmt = select(ScoreRecord).where(ScoreRecord.course_id == course_id)
        if student_ids:
            stmt = stmt.where(ScoreRecord.student_id.in_(student_ids))  # type: ignore[arg-type]
        if batch_id:
            stmt = stmt.where(ScoreRecord.batch_id == batch_id)
        if keyword and not student_ids:
            stmt = stmt.where(ScoreRecord.student_id == -1)

        for s in session.exec(stmt).all():
            student = session.get(Student, s.student_id)
            batch = session.get(ExamBatch, s.batch_id)
            if not student:
                continue
            rows.append({
                "id": f"score_{s.score_id}",
                "recordId": s.score_id,
                "dataType": "score",
                "studentId": student.student_no,
                "studentName": student.real_name,
                "courseId": course_id,
                "courseName": "",
                "semester": batch.semester if batch else "",
                "batchId": s.batch_id,
                "batchName": batch.batch_name if batch else "",
                "score": s.score,
                "isPass": s.is_pass,
                "remark": s.remark,
                "sourceData": s.source_data,
            })

        # 新表：IndividualScore
        istmt = select(IndividualScore).join(
            ExamBatch, IndividualScore.exam_batch_id == ExamBatch.batch_id
        ).where(ExamBatch.course_id == course_id)
        if student_ids:
            istmt = istmt.where(IndividualScore.student_id.in_(student_ids))  # type: ignore[arg-type]
        if batch_id:
            istmt = istmt.where(IndividualScore.exam_batch_id == batch_id)
        if keyword and not student_ids:
            istmt = istmt.where(IndividualScore.student_id == -1)

        for s in session.exec(istmt).all():
            student = session.get(Student, s.student_id)
            batch = session.get(ExamBatch, s.exam_batch_id)
            if not student:
                continue
            rows.append({
                "id": f"individual_{s.score_id}",
                "recordId": s.score_id,
                "dataType": "score",
                "subType": "individual_score",
                "studentId": student.student_no,
                "studentName": student.real_name,
                "courseId": course_id,
                "courseName": "",
                "semester": batch.semester if batch else "",
                "batchId": s.exam_batch_id,
                "batchName": batch.batch_name if batch else "",
                "score": s.score,
                "sourceData": s.source_data,
            })

        # 新表：CourseTestDetail
        ctstmt = select(CourseTestDetail).join(
            ExamBatch, CourseTestDetail.exam_batch_id == ExamBatch.batch_id
        ).where(ExamBatch.course_id == course_id)
        if student_ids:
            ctstmt = ctstmt.where(CourseTestDetail.student_id.in_(student_ids))  # type: ignore[arg-type]
        if batch_id:
            ctstmt = ctstmt.where(CourseTestDetail.exam_batch_id == batch_id)
        if keyword and not student_ids:
            ctstmt = ctstmt.where(CourseTestDetail.student_id == -1)

        for s in session.exec(ctstmt).all():
            student = session.get(Student, s.student_id)
            batch = session.get(ExamBatch, s.exam_batch_id)
            if not student:
                continue
            rows.append({
                "id": f"test_detail_{s.score_id}",
                "recordId": s.score_id,
                "dataType": "score",
                "subType": "course_test_detail",
                "studentId": student.student_no,
                "studentName": student.real_name,
                "courseId": course_id,
                "courseName": "",
                "semester": batch.semester if batch else "",
                "batchId": s.exam_batch_id,
                "batchName": batch.batch_name if batch else "",
                "score": s.total_score,
                "totalScore": s.total_score,
                "questionScores": [
                    s.question1_score, s.question2_score,
                    s.question3_score, s.question4_score, s.question5_score,
                ],
                "knowledgePoints": [
                    s.question1_knowledge, s.question2_knowledge,
                    s.question3_knowledge, s.question4_knowledge, s.question5_knowledge,
                ],
                "sourceData": s.source_data,
            })

    # ── 考勤数据 ──
    if data_type is None or data_type == "attendance":
        # 旧表：AttendanceRecord
        stmt = select(AttendanceRecord).where(AttendanceRecord.course_id == course_id)
        if student_ids:
            stmt = stmt.where(AttendanceRecord.student_id.in_(student_ids))  # type: ignore[arg-type]
        if keyword and not student_ids:
            stmt = stmt.where(AttendanceRecord.student_id == -1)

        for a in session.exec(stmt).all():
            student = session.get(Student, a.student_id)
            if not student:
                continue
            rows.append({
                "id": f"attendance_{a.attendance_id}",
                "recordId": a.attendance_id,
                "dataType": "attendance",
                "studentId": student.student_no,
                "studentName": student.real_name,
                "courseId": course_id,
                "courseName": "",
                "semester": "",
                "attendanceDate": a.attendance_date.isoformat() if a.attendance_date else None,
                "status": STATUS_MAP.get(a.status, "未知"),
                "statusCode": a.status,
                "remark": a.remark,
                "sourceData": a.source_data,
            })

        # 新表：AttendanceSheet
        ash_stmt = select(AttendanceSheet).join(
            ExamBatch, AttendanceSheet.exam_batch_id == ExamBatch.batch_id
        ).where(ExamBatch.course_id == course_id)
        if student_ids:
            ash_stmt = ash_stmt.where(AttendanceSheet.student_id.in_(student_ids))  # type: ignore[arg-type]
        if keyword and not student_ids:
            ash_stmt = ash_stmt.where(AttendanceSheet.student_id == -1)

        for a in session.exec(ash_stmt).all():
            student = session.get(Student, a.student_id)
            batch = session.get(ExamBatch, a.exam_batch_id)
            if not student:
                continue
            att_data = [
                a.attendance_1, a.attendance_2, a.attendance_3, a.attendance_4,
                a.attendance_5, a.attendance_6, a.attendance_7, a.attendance_8,
                a.attendance_9, a.attendance_10, a.attendance_11, a.attendance_12,
                a.attendance_13, a.attendance_14, a.attendance_15, a.attendance_16,
                a.attendance_17, a.attendance_18, a.attendance_19, a.attendance_20,
                a.attendance_21, a.attendance_22, a.attendance_23, a.attendance_24,
                a.attendance_25, a.attendance_26, a.attendance_27, a.attendance_28,
                a.attendance_29, a.attendance_30, a.attendance_31, a.attendance_32,
            ]
            rows.append({
                "id": f"attendance_sheet_{a.score_id}",
                "recordId": a.score_id,
                "dataType": "attendance",
                "subType": "attendance_sheet",
                "studentId": student.student_no,
                "studentName": student.real_name,
                "courseId": course_id,
                "courseName": "",
                "semester": batch.semester if batch else "",
                "batchId": a.exam_batch_id,
                "batchName": batch.batch_name if batch else "",
                "attendanceData": att_data,
                "totalCount": a.total_count,
                "presentCount": a.present_count,
                "leaveCount": a.leave_count,
                "lateCount": a.late_count,
                "earlyLeaveCount": a.early_leave_count,
                "attendanceRate": a.attendance_rate,
                "sourceData": a.source_data,
            })

    # 分页
    total = len(rows)
    start = (page - 1) * page_size
    end = start + page_size
    paged = rows[start:end]

    return {
        "total": total,
        "page": page,
        "pageSize": page_size,
        "totalPages": (total + page_size - 1) // page_size if total else 0,
        "data": paged,
    }


# ============================================================================
# 编辑接口（3.2.7 节：Data.Query.Edit）
# ============================================================================


# NOTE: /{record_id}/row 必须放在 /{record_type}/{record_id} 之前，
# 否则旧路由会错误匹配 "112/row" → record_type=112, record_id="row"
@router.put("/teaching-data/{record_id}/row", tags=["教学数据"])
def update_row_data(
    record_id: int,
    payload: Any = Body(...),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """更新一条教学数据的完整行信息（含所有字段和子记录）。

    传入完整的 source_data JSON 对象，后端自动：
      - 更新主记录的 score / status 和 source_data
      - 对于考试扣分类型，删除旧的各题子记录，按新数据重建
    """
    src = payload.get("source_data") if isinstance(payload, dict) else None

    record = session.get(ScoreRecord, record_id)
    if record is None:
        record = session.get(AttendanceRecord, record_id)
    if record is None:
        record = session.get(IndividualScore, record_id)
    if record is None:
        record = session.get(CourseTestDetail, record_id)
    if record is None:
        record = session.get(AttendanceSheet, record_id)
    if record is None:
        raise HTTPException(status_code=404, detail="记录不存在")

    if isinstance(record, ScoreRecord):
        course_id = record.course_id
        _require_teacher_for_course(current_user, course_id, session)

        if src is not None:
            # 更新主记录分数
            total = src.get("总成绩")
            if total is not None:
                try:
                    record.score = float(total)
                    record.is_pass = 1 if record.score >= 60 else 0
                except (ValueError, TypeError):
                    pass
            # 更新 source_data
            record.source_data = json.dumps(src, ensure_ascii=False)
            record.update_time = datetime.now()
            session.add(record)

            # 对于考试扣分类型：删除旧子记录，按新数据重建各题扣分
            batch = session.get(ExamBatch, record.batch_id)
            if batch and "大题" not in (batch.batch_name or ""):
                _rebuild_question_sub_records(session, record, src)

            session.commit()

        return {"recordId": record_id, "recordType": "score", "updated": True}

    elif isinstance(record, AttendanceRecord):
        course_id = record.course_id
        _require_teacher_for_course(current_user, course_id, session)

        if src is not None:
            record.source_data = json.dumps(src, ensure_ascii=False)
            record.update_time = datetime.now()
            session.add(record)
            session.commit()

        return {"recordId": record_id, "recordType": "attendance", "updated": True}

    elif isinstance(record, IndividualScore):
        batch = session.get(ExamBatch, record.exam_batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="关联考试批次不存在")
        _require_teacher_for_course(current_user, batch.course_id, session)

        if src is not None:
            score_val = src.get("成绩")
            if score_val is not None:
                try:
                    record.score = float(score_val)
                except (ValueError, TypeError):
                    pass
            record.source_data = json.dumps(src, ensure_ascii=False)
            record.update_time = datetime.now()
            session.add(record)
            session.commit()

        return {"recordId": record_id, "recordType": "individual_score", "updated": True}

    elif isinstance(record, CourseTestDetail):
        batch = session.get(ExamBatch, record.exam_batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="关联考试批次不存在")
        _require_teacher_for_course(current_user, batch.course_id, session)

        if src is not None:
            total = src.get("总成绩")
            if total is not None:
                try:
                    record.total_score = float(total)
                except (ValueError, TypeError):
                    pass
            record.source_data = json.dumps(src, ensure_ascii=False)
            record.update_time = datetime.now()
            session.add(record)
            session.commit()

        return {"recordId": record_id, "recordType": "course_test_detail", "updated": True}

    elif isinstance(record, AttendanceSheet):
        batch = session.get(ExamBatch, record.exam_batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="关联考试批次不存在")
        _require_teacher_for_course(current_user, batch.course_id, session)

        if src is not None:
            record.source_data = json.dumps(src, ensure_ascii=False)
            record.update_time = datetime.now()
            session.add(record)
            session.commit()

        return {"recordId": record_id, "recordType": "attendance_sheet", "updated": True}

    raise HTTPException(status_code=400, detail="记录类型不支持")


def _rebuild_question_sub_records(
    session: Session,
    main_record: ScoreRecord,
    src: dict,
) -> None:
    """根据完整行数据重建考试扣分各题子记录。"""
    from app.models.exam import ScoreRecord as SR, ExamBatch as EB

    exam_name = str(src.get("测试名称") or "考试").strip()
    student_id = main_record.student_id
    course_id = main_record.course_id
    create_by = main_record.create_by

    for qn in range(1, 6):
        q_batch_name = f"{exam_name}-第{qn}大题"
        # 查找已有子批次
        existing_batch = session.exec(
            select(EB).where(EB.course_id == course_id, EB.batch_name == q_batch_name)
        ).first()

        deduction_raw = src.get(f"第{qn}大题")
        deduction = None
        if deduction_raw is not None and str(deduction_raw).strip() != "":
            try:
                deduction = float(deduction_raw)
            except (ValueError, TypeError):
                deduction = None

        # 删除旧子记录
        if existing_batch:
            old_records = session.exec(
                select(SR).where(
                    SR.student_id == student_id,
                    SR.batch_id == existing_batch.batch_id,
                )
            ).all()
            for old in old_records:
                session.delete(old)

        # 如果有扣分数据，创建新子记录
        if deduction is not None and deduction != 0.0 and existing_batch:
            knowledge = str(src.get(f"第 {qn} 大题扣分的主要知识点", "")).strip()
            session.add(SR(
                course_id=course_id,
                student_id=student_id,
                batch_id=existing_batch.batch_id,
                score=float(deduction),
                is_pass=1,
                remark=knowledge,
                source_data=main_record.source_data,
                create_by=create_by,
            ))
    session.commit()


@router.put("/teaching-data/{record_type}/{record_id}", tags=["教学数据"])
def edit_teaching_data(
    record_type: str,
    record_id: int,
    score: float | None = Query(default=None, description="成绩值（score/individual_score 类型）"),
    status_code: int | None = Query(default=None, description="考勤状态: 0=出勤 1=迟到 2=早退 3=缺勤 4=请假"),
    remark: str | None = Query(default=None, description="备注"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """编辑单条教学数据（Data.Query.Edit）。

    支持修改成绩值、考勤状态、备注。仅授课教师可修改自己课程的数据。
    支持旧表（ScoreRecord/AttendanceRecord）和新表（IndividualScore/AttendanceSheet/CourseTestDetail）。
    """
    if record_type == "score":
        record = session.get(ScoreRecord, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="成绩记录不存在")
        course_id = record.course_id
        _require_teacher_for_course(current_user, course_id, session)

        if score is not None:
            record.score = score
            record.is_pass = 1 if score >= 60 else 0
        if remark is not None:
            record.remark = remark
        record.update_time = datetime.now()
        session.add(record)
        session.commit()
        session.refresh(record)
        return {
            "recordType": "score",
            "recordId": record.score_id,
            "score": record.score,
            "isPass": record.is_pass,
            "remark": record.remark,
        }

    elif record_type == "individual_score":
        record = session.get(IndividualScore, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="单项成绩记录不存在")
        batch = session.get(ExamBatch, record.exam_batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="关联考试批次不存在")
        _require_teacher_for_course(current_user, batch.course_id, session)

        if score is not None:
            record.score = score
        record.update_time = datetime.now()
        session.add(record)
        session.commit()
        session.refresh(record)
        return {
            "recordType": "individual_score",
            "recordId": record.score_id,
            "score": record.score,
        }

    elif record_type == "course_test_detail":
        record = session.get(CourseTestDetail, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="课程测试记录不存在")
        batch = session.get(ExamBatch, record.exam_batch_id)
        if not batch:
            raise HTTPException(status_code=404, detail="关联考试批次不存在")
        _require_teacher_for_course(current_user, batch.course_id, session)

        if score is not None:
            record.total_score = score
        record.update_time = datetime.now()
        session.add(record)
        session.commit()
        session.refresh(record)
        return {
            "recordType": "course_test_detail",
            "recordId": record.score_id,
            "totalScore": record.total_score,
        }

    elif record_type == "attendance":
        record = session.get(AttendanceRecord, record_id)
        if not record:
            raise HTTPException(status_code=404, detail="考勤记录不存在")
        course_id = record.course_id
        _require_teacher_for_course(current_user, course_id, session)

        if status_code is not None:
            if status_code not in STATUS_MAP:
                raise HTTPException(status_code=400, detail=f"无效的考勤状态: {status_code}，有效值: 0-4")
            record.status = status_code
        if remark is not None:
            record.remark = remark
        record.update_time = datetime.now()
        session.add(record)
        session.commit()
        session.refresh(record)
        return {
            "recordType": "attendance",
            "recordId": record.attendance_id,
            "status": STATUS_MAP.get(record.status, "未知"),
            "statusCode": record.status,
            "remark": record.remark,
        }

    else:
        raise HTTPException(status_code=400, detail=f"不支持的数据类型: {record_type}，支持: score, individual_score, course_test_detail, attendance")


@router.get("/teaching-data/export", tags=["教学数据"])
def export_teaching_data(
    course_id: int = Query(..., description="课程 ID"),
    keyword: str | None = Query(default=None, description="学生姓名/学号模糊搜索"),
    data_type: str | None = Query(default=None, description="数据类型: score / attendance"),
    batch_id: int | None = Query(default=None, description="考核批次 ID"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> Response:
    """将查询结果导出为 Excel 文件（Data.Query.Export）。

    权限：仅授课教师可导出自己课程的数据。
    """
    _require_teacher_for_course(current_user, course_id, session)
    course = session.get(Course, course_id)

    # 复用查询逻辑获取数据
    result = query_teaching_data(
        course_id=course_id,
        keyword=keyword,
        data_type=data_type,
        batch_id=batch_id,
        page=1,
        page_size=10000,  # 导出全部
        session=session,
        current_user=current_user,
    )
    rows: list[dict] = result["data"]

    # 生成 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "教学数据"

    # 表头样式
    header_font = openpyxl.styles.Font(bold=True, size=11)
    header_fill = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    if not rows:
        ws.cell(row=1, column=1, value="无匹配数据").font = header_font
    else:
        # 根据第一条决定列
        if rows[0]["dataType"] == "score":
            headers = ["序号", "学号", "姓名", "数据类型", "考核批次", "成绩", "是否及格", "备注"]
        else:
            headers = ["序号", "学号", "姓名", "数据类型", "考勤日期", "出勤状态", "备注"]

        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for ri, row in enumerate(rows, start=2):
            if row["dataType"] == "score":
                vals = [
                    ri - 1, row["studentId"], row["studentName"], "成绩",
                    row.get("batchName", ""), row.get("score", ""),
                    "是" if row.get("isPass") else "否", row.get("remark", ""),
                ]
            else:
                vals = [
                    ri - 1, row["studentId"], row["studentName"], "考勤",
                    row.get("attendanceDate", ""), row.get("status", ""),
                    row.get("remark", ""),
                ]
            for ci, v in enumerate(vals, start=1):
                ws.cell(row=ri, column=ci, value=v)

    # 调整列宽
    for ci in range(1, 15):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 16

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = f"{course.course_name}_教学数据.xlsx" if course else "教学数据导出.xlsx"
    encoded = quote(safe_name, safe="")
    ascii_name = "teaching_data_export.xlsx"

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{ascii_name}\"; "
                f"filename*=UTF-8''{encoded}"
            ),
        },
    )


# ============================================================================
# 文件上传导入接口
# ============================================================================


@router.post("/teaching-data/upload", tags=["教学数据"])
def upload_teaching_data(
    file: UploadFile = File(...),
    course_id: int = Query(..., description="课程 ID"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> dict:
    """上传教学数据文件并批量导入。

    权限：必须登录，且为对应课程的任课教师。
    格式：仅 .xlsx / .txt（UTF-8 逗号分隔）。
    模板：自动检测匹配课程测试各题扣分情况 / 成绩汇总 / 成绩考勤情况。
    """
    # ------ 1. 登录验证（Data.FileUpload.UserValid）------
    if not current_user:
        raise HTTPException(status_code=401, detail="请先登录")

    # ------ 2. 文件格式校验（Data.FileUpload.Format）------
    _, ext = os.path.splitext(file.filename or "")
    ext = ext.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"仅支持 .xlsx 和 UTF-8 逗号分隔 .txt 格式，当前文件扩展名为「{ext}」",
        )

    # ------ 3. 课程与权限校验（Data.FileUpload.UserValid.Logined）------
    course = session.get(Course, course_id)
    if not course:
        raise HTTPException(status_code=404, detail="课程不存在")

    # 查找当前用户对应的教师记录
    teacher = session.exec(
        select(Teacher).where(Teacher.user_id == current_user.user_id)
    ).first()
    if not teacher:
        raise HTTPException(
            status_code=403,
            detail="仅任课教师可上传教学数据，当前账号未关联教师信息",
        )
    if course.teacher_id != teacher.teacher_id:
        raise HTTPException(
            status_code=403,
            detail=f"仅授课教师可上传数据。课程「{course.course_name}」的授课教师与当前账号不匹配",
        )

    # ------ 4. 保存临时文件 & 导入 ------
    tmp_path = os.path.join(
        tempfile.gettempdir(),
        f"teaching_data_upload_{uuid.uuid4().hex}{ext}",
    )
    try:
        content = file.file.read()
        # 对于 .txt 文件，校验 UTF-8 编码
        if ext == ".txt":
            try:
                content.decode("utf-8")
            except UnicodeDecodeError:
                raise HTTPException(
                    status_code=400,
                    detail="Txt 文件必须是 UTF-8 编码，当前文件编码不符",
                )

        with open(tmp_path, "wb") as f:
            f.write(content)

        # 调用导入服务
        result: ImportResult = import_file(
            session=session,
            file_path=tmp_path,
            file_ext=ext,
            course_id=course_id,
            create_by=current_user.user_id,
        )

        # 导入成功后自动刷新课程分析数据
        # （学情画像、知识点掌握度、学习质量评价、学情预警）
        analysis_refresh: dict = {}
        if result.success_count > 0:
            try:
                analysis_refresh = refresh_course_analysis(session, course_id)
            except Exception as e:
                analysis_refresh = {"error": str(e)}

    finally:
        # 清理临时文件
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    # ------ 5. 返回结果（Data.FileUpload.Result）------
    return {
        "fileName": file.filename,
        "courseId": course_id,
        "courseName": course.course_name,
        "detectedTemplate": result.detected_template,
        "sheetsProcessed": result.sheets_processed,
        "successCount": result.success_count,
        "errorCount": result.error_count,
        "errors": [
            {
                "sheet": e.sheet,
                "row": e.row,
                "field": e.field,
                "message": e.message,
            }
            for e in result.errors
        ],
        "analysisRefresh": analysis_refresh,
    }


# ============================================================================
# 模板下载接口（3.2.5 节）
# ============================================================================

def _check_template_access(current_user: SysUser, session: Session) -> None:
    """校验模板下载权限：仅任课教师（Data.Template.UserValid）。"""
    role = session.get(SysRole, current_user.role_id)
    role_code = role.role_code if role else ""
    if role_code != "teacher":
        raise HTTPException(
            status_code=403,
            detail="仅任课教师可下载模板",
        )


@router.get("/teaching-data/templates", tags=["教学数据"])
def list_templates(
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> list[dict]:
    """返回可下载的模板列表（Data.Template.Type）。

    按 dataType 分组：成绩、考勤。
    权限：任课教师。
    """
    _check_template_access(current_user, session)
    return [
        {
            "templateId": m["template_id"],
            "name": m["name"],
            "dataType": m["dataType"],
            "description": m["description"],
            "headers": m["headers"],
        }
        for m in TEMPLATE_META
    ]


@router.get("/teaching-data/templates/{template_id}", tags=["教学数据"])
def download_template(
    template_id: str,
    format: str = Query(default="xlsx", description="下载格式: xlsx / txt"),
    session: Session = Depends(get_session),
    current_user: SysUser = Depends(get_current_user),
) -> Response:
    """下载指定模板文件（Data.Template.Download）。

    支持两种格式：
      - xlsx（默认）：Excel 格式，含表头样式和填写说明
      - txt：UTF-8 逗号分隔文本，含注释说明、表头行和示例行

    权限：任课教师。
    """
    _check_template_access(current_user, session)

    fmt = format.strip().lower()
    if fmt not in ("xlsx", "txt"):
        raise HTTPException(status_code=400, detail=f"不支持的格式: {format}，仅支持 xlsx 和 txt")

    meta = next((m for m in TEMPLATE_META if m["template_id"] == template_id), None)
    if not meta:
        raise HTTPException(status_code=404, detail=f"模板不存在: {template_id}")

    try:
        if fmt == "txt":
            file_bytes = generate_template_txt(template_id)
            media_type = "text/plain; charset=utf-8"
            safe_name = f"模板-{meta['name']}.txt"
            ascii_name = f"template-{template_id}.txt"
        else:
            file_bytes = generate_template_xlsx(template_id)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            safe_name = f"模板-{meta['name']}.xlsx"
            ascii_name = f"template-{template_id}.xlsx"
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    # 文件名含中文，用 RFC 5987 编码
    encoded = quote(safe_name, safe="")
    return Response(
        content=file_bytes,
        media_type=media_type,
        headers={
            "Content-Disposition": (
                f"attachment; filename=\"{ascii_name}\"; "
                f"filename*=UTF-8''{encoded}"
            ),
        },
    )
