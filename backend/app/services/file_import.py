"""文件导入服务：检测模板、校验数据、批量导入教学数据。

支持的模板：
  1. 课程测试各题扣分情况 → ExamBatch + CourseTestDetail + 知识点自动创建
  2. 单项成绩             → ExamBatch + IndividualScore
  3. 成绩考勤情况         → ExamBatch + AttendanceSheet
"""

from __future__ import annotations

import csv
import io as _io
import json
import logging
import os
import re
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import openpyxl
from sqlmodel import Session, select

from app.models import (
    AttendanceSheet,
    Course,
    CourseStudent,
    CourseTestDetail,
    ExamBatch,
    IndividualScore,
    KnowledgeModule,
    KnowledgePoint,
    Student,
    SysUser,
)


# ============================================================================
# 模板定义
# ============================================================================

@dataclass
class TemplateDef:
    """模板定义：用于匹配和导入。"""

    template_id: str
    name: str
    # 用于匹配的关键表头（只要这些字段全部存在，即命中模板）
    match_headers: list[str]
    # 所有可能的表头（包括可选字段）
    all_headers: list[str]
    # 必填字段列表（行数据中必须非空）
    required_fields: list[str]
    # 数字字段列表（用于类型校验）
    numeric_fields: list[str]


# 模板1：课程测试各题扣分情况
TEMPLATE_EXAM_DEDUCTION = TemplateDef(
    template_id="exam_deduction",
    name="课程测试各题扣分情况",
    match_headers=["学号", "第1大题", "第2大题", "总成绩"],
    all_headers=[
        "编号", "课程号", "课程名称", "学期", "测试名称", "学号", "姓名",
        "第1大题", "第 1 大题扣分的主要知识点",
        "第2大题", "第 2 大题扣分的主要知识点",
        "第3大题", "第 3 大题扣分的主要知识点",
        "第4大题", "第 4 大题扣分的主要知识点",
        "第5大题", "第 5 大题扣分的主要知识点",
        "总成绩",
    ],
    required_fields=["学号", "姓名"],
    numeric_fields=["第1大题", "第2大题", "第3大题", "第4大题", "第5大题", "总成绩"],
)

# 模板2：单项成绩
TEMPLATE_SIMPLE_SCORE = TemplateDef(
    template_id="simple_score",
    name="单项成绩",
    match_headers=["学号", "成绩名称", "成绩"],
    all_headers=[
        "编号", "课程号", "课程名称", "学期", "学号", "姓名",
        "成绩名称", "成绩",
    ],
    required_fields=["学号", "姓名"],
    numeric_fields=["成绩"],
)

# 模板3：成绩考勤情况
TEMPLATE_ATTENDANCE = TemplateDef(
    template_id="attendance",
    name="成绩考勤情况",
    match_headers=["学号", "考勤1", "考勤2", "考勤3"],
    all_headers=[
        "编号", "课程号", "课程名称", "学期", "学号", "姓名",
        *[f"考勤{i}" for i in range(1, 33)],
        "考勤总数", "到课数", "请假数", "早退数", "迟到数", "到课率",
    ],
    required_fields=["学号", "姓名"],
    numeric_fields=[],
)

ALL_TEMPLATES: list[TemplateDef] = [
    TEMPLATE_EXAM_DEDUCTION,
    TEMPLATE_SIMPLE_SCORE,
    TEMPLATE_ATTENDANCE,
]


# ============================================================================
# 结果与错误结构
# ============================================================================

@dataclass
class ImportError:
    """单条导入错误。"""

    sheet: str
    row: int
    field: str | None
    message: str


@dataclass
class ImportResult:
    """导入结果汇总。"""

    success_count: int = 0
    error_count: int = 0
    errors: list[ImportError] = field(default_factory=list)
    detected_template: str | None = None
    sheets_processed: list[str] = field(default_factory=list)


# ============================================================================
# 工具函数
# ============================================================================

def _is_formula(value: Any) -> bool:
    """判断是否为 Excel 公式（字符串以 = 开头）。"""
    return isinstance(value, str) and value.startswith("=")


def _safe_float(value: Any) -> float | None:
    """安全转换为 float，公式或非法值返回 None。"""
    if value is None:
        return None
    if _is_formula(value):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _safe_int(value: Any) -> int | None:
    """安全转换为 int。"""
    f = _safe_float(value)
    if f is not None:
        return int(f)
    return None


def _safe_str(value: Any) -> str | None:
    """安全转换为字符串。"""
    if value is None:
        return None
    if _is_formula(value):
        return None
    return str(value).strip()


def _normalize_headers(headers: list[str | None]) -> list[str]:
    """过滤 None 并去除首尾空格。"""
    return [h.strip() for h in headers if h is not None and str(h).strip()]


# ============================================================================
# 模板检测
# ============================================================================

def detect_template(headers: list[str]) -> TemplateDef | None:
    """根据表头检测匹配的模板。

    匹配逻辑：模板的 match_headers 全部存在于文件表头中即视为命中。
    若多个模板命中，取匹配数量最多的；平局时取第一个。
    """
    header_set = set(headers)
    best: tuple[TemplateDef, int] | None = None

    for tmpl in ALL_TEMPLATES:
        match_count = sum(1 for h in tmpl.match_headers if h in header_set)
        if match_count == len(tmpl.match_headers):
            if best is None or match_count > best[1]:
                best = (tmpl, match_count)

    return best[0] if best else None


# ============================================================================
# 行级校验
# ============================================================================

def validate_row(
    row_data: dict[str, Any],
    tmpl: TemplateDef,
    sheet_name: str,
    row_index: int,  # 1-based 数据行号（Excel 行号）
) -> list[ImportError]:
    """校验单行数据，返回错误列表。"""
    errors: list[ImportError] = []

    # 必填字段校验
    for field in tmpl.required_fields:
        if field not in row_data or row_data[field] is None or str(row_data[field]).strip() == "":
            errors.append(ImportError(
                sheet=sheet_name,
                row=row_index,
                field=field,
                message=f"必填字段「{field}」为空",
            ))

    # 数字字段校验
    for field in tmpl.numeric_fields:
        if field in row_data and row_data[field] is not None:
            val = row_data[field]
            if _is_formula(val):
                errors.append(ImportError(
                    sheet=sheet_name,
                    row=row_index,
                    field=field,
                    message=f"字段「{field}」包含公式，请上传前将公式转为数值",
                ))
            elif not isinstance(val, (int, float)):
                try:
                    float(val)
                except (ValueError, TypeError):
                    errors.append(ImportError(
                        sheet=sheet_name,
                        row=row_index,
                        field=field,
                        message=f"字段「{field}」应为数字，实际值为「{val}」",
                    ))

    # 学号字段校验（应为数字串或纯数字）
    student_no = str(row_data.get("学号", "")).strip()
    if student_no and not student_no.isdigit():
        errors.append(ImportError(
            sheet=sheet_name,
            row=row_index,
            field="学号",
            message=f"学号应为纯数字，实际值为「{student_no}」",
        ))

    return errors


# ============================================================================
# 文件解析
# ============================================================================

def _find_header_row(
    rows: list[tuple[Any, ...]],
) -> tuple[int, list[str], dict[int, str]]:
    """在 sheet 的前几行中寻找表头行。

    返回 (header_row_index, headers_list, col_index_to_name)。
    若找不到则 header_row_index=-1。
    """
    for ri in range(min(4, len(rows))):
        raw_headers = [str(c).strip() if c is not None else None for c in rows[ri]]
        headers = [h for h in raw_headers if h]
        if len(headers) >= 2:
            # 构建列索引 → 列名映射
            col_map: dict[int, str] = {}
            for ci, h in enumerate(raw_headers):
                if h:
                    col_map[ci] = h
            return ri, headers, col_map
    return -1, [], {}


def parse_xlsx(file_path: str) -> dict[str, list[dict[str, Any]]]:
    """解析 .xlsx 文件，返回 {sheet_name: [row_dict, ...], ...}。

    自动在前 4 行中寻找表头行（不再硬编码第1行）。
    忽略全空行；公式单元格保留字符串形式。
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx, headers, col_map = _find_header_row(rows)
        if header_idx < 0 or not headers:
            continue

        sheet_data: list[dict[str, Any]] = []
        for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
            row_dict: dict[str, Any] = {}
            all_none = True
            for col_idx, cell_value in enumerate(row):
                if col_idx in col_map:
                    header_name = col_map[col_idx]
                    row_dict[header_name] = cell_value
                if cell_value is not None:
                    all_none = False

            # 跳过全空行
            if all_none or not row_dict:
                continue
            # 跳过无身份标识的汇总/统计行：学号和姓名同时为空
            sno_val = row_dict.get("学号")
            name_val = row_dict.get("姓名")
            if (sno_val is None or str(sno_val).strip() == "") and \
               (name_val is None or str(name_val).strip() == ""):
                continue

            sheet_data.append({"_excel_row": row_idx, **row_dict})

        if sheet_data:
            result[sheet_name] = sheet_data

    wb.close()
    return result


def parse_txt(file_path: str) -> dict[str, list[dict[str, Any]]]:
    """解析逗号分隔的 UTF-8 .txt 文件，返回 {"default": [row_dict, ...]}。"""
    result: dict[str, list[dict[str, Any]]] = {}
    with open(file_path, encoding="utf-8") as f:
        content = f.read()

    # 使用 csv 模块解析
    reader = csv.reader(_io.StringIO(content))
    rows = list(reader)
    if not rows:
        return result

    headers = _normalize_headers(rows[0])
    if not headers:
        return result

    sheet_data: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        row_dict: dict[str, Any] = {}
        all_empty = True
        for i, cell_value in enumerate(row):
            if i < len(headers):
                header_name = headers[i]
                val = cell_value.strip() if cell_value else None
                row_dict[header_name] = val
                if val:
                    all_empty = False

        # 跳过全空行
        if all_empty or not row_dict:
            continue
        # 跳过无身份标识的汇总/统计行：学号和姓名同时为空
        sno_val = row_dict.get("学号")
        name_val = row_dict.get("姓名")
        if (sno_val is None or str(sno_val).strip() == "") and \
           (name_val is None or str(name_val).strip() == ""):
            continue

        sheet_data.append({"_excel_row": row_idx, **row_dict})

    result["default"] = sheet_data
    return result


# ============================================================================
# 数据导入辅助
# ============================================================================

def _get_or_create_exam_batch(
    session: Session,
    course_id: int,
    batch_name: str,
    batch_type: int,
    semester: str,
    create_by: int,
    batch_weight: float | None = None,
) -> ExamBatch:
    """获取或创建考核批次（按 course_id + batch_name + semester 去重）。"""
    stmt = select(ExamBatch).where(
        ExamBatch.course_id == course_id,
        ExamBatch.batch_name == batch_name,
        ExamBatch.semester == semester,
    )
    batch = session.exec(stmt).first()
    if batch:
        # 如果已有批次权重为空且有默认权重，补填权重
        if batch.batch_weight is None and batch_weight is not None:
            batch.batch_weight = batch_weight
            session.add(batch)
            session.commit()
        return batch

    batch = ExamBatch(
        course_id=course_id,
        batch_name=batch_name,
        batch_type=batch_type,
        semester=semester,
        batch_weight=batch_weight,
        full_score=100,
        create_by=create_by,
    )
    session.add(batch)
    session.commit()
    session.refresh(batch)
    return batch


def _get_student_by_no(session: Session, student_no: str) -> Student | None:
    """根据学号查找学生。"""
    return session.exec(
        select(Student).where(Student.student_no == str(student_no).strip())
    ).first()


def _ensure_course_student(session: Session, course_id: int, student_id: int) -> None:
    """确保学生在课程选修表中存在。"""
    exists = session.exec(
        select(CourseStudent).where(
            CourseStudent.course_id == course_id,
            CourseStudent.student_id == student_id,
        )
    ).first()
    if not exists:
        session.add(CourseStudent(
            course_id=course_id,
            student_id=student_id,
            status=1,
        ))
        session.commit()


def _extract_semester(rows: list[dict[str, Any]]) -> str:
    """从数据行中提取学期（取第一个非空值）。"""
    for row in rows:
        sem = row.get("学期")
        if sem and str(sem).strip():
            return str(sem).strip()
    return ""


def _map_batch_type_by_name(batch_name: str) -> int:
    """根据成绩名称推断 batch_type。

    映射规则：
      期中 → 3, 期末 → 4, 实验 → 2, 作业/平时 → 1, 小班讨论 → 1, 默认 → 1
    """
    name = batch_name.strip()
    if "期中" in name:
        return 3
    if "期末" in name:
        return 4
    if "实验" in name:
        return 2
    if "小班讨论" in name or "讨论" in name:
        return 1
    return 1  # 平时成绩


def _get_default_batch_weight(batch_name: str, batch_type: int) -> float | None:
    """根据批次名称/类型返回默认权重。

    期中考试 → 35, 期末考试 → 50, 考勤 → 5, 小班讨论 → 10
    其他类型返回 None，表示不参与加权计算。
    """
    name = batch_name.strip()
    if batch_type == 5:
        return 5.0       # 考勤
    if "小班讨论" in name or "讨论" in name:
        return 10.0
    if "期末" in name:
        return 50.0
    if "期中" in name:
        return 35.0
    return None


# ============================================================================
# 知识点自动管理
# ============================================================================

def _ensure_knowledge_point(
    session: Session,
    course_id: int,
    point_name: str,
) -> KnowledgePoint:
    """获取或创建知识点（按名称，课程范围内）。

    若课程下无知识模块，自动创建"默认模块"。
    知识点按名称查找，有则不变，无则新增。
    """
    # 1. 查找课程下所有模块
    modules = session.exec(
        select(KnowledgeModule).where(KnowledgeModule.course_id == course_id)
    ).all()

    if not modules:
        # 自动创建默认模块
        default_module = KnowledgeModule(
            course_id=course_id,
            module_name="默认模块",
            description="自动创建的知识模块（导入数据时生成）",
            sort_num=0,
        )
        session.add(default_module)
        session.commit()
        session.refresh(default_module)
        modules = [default_module]

    # 2. 获取所有模块 ID
    module_ids = [m.module_id for m in modules]

    # 3. 查找是否已存在同名知识点
    existing = session.exec(
        select(KnowledgePoint).where(
            KnowledgePoint.module_id.in_(module_ids),  # type: ignore[arg-type]
            KnowledgePoint.point_name == point_name,
        )
    ).first()

    if existing:
        return existing

    # 4. 不存在则新增到第一个模块下
    point = KnowledgePoint(
        module_id=modules[0].module_id,
        point_name=point_name,
        description="",
        sort_num=0,
    )
    session.add(point)
    session.commit()
    session.refresh(point)
    return point


def _collect_and_ensure_knowledge_points(
    session: Session,
    course_id: int,
    rows: list[dict[str, Any]],
) -> dict[str, KnowledgePoint]:
    """从课程测试数据行中收集所有知识点名称并确保其存在。

    返回 {知识点名称: KnowledgePoint} 映射。
    """
    knowledge_names: set[str] = set()
    for row_data in rows:
        for qn in range(1, 6):
            col_name = f"第 {qn} 大题扣分的主要知识点"
            kn_str = str(row_data.get(col_name, "")).strip() if row_data.get(col_name) else ""
            if kn_str:
                knowledge_names.add(kn_str)

    result: dict[str, KnowledgePoint] = {}
    for name in knowledge_names:
        result[name] = _ensure_knowledge_point(session, course_id, name)

    return result


# ============================================================================
# 模板1：课程测试各题扣分情况
# ============================================================================

def _import_exam_deduction(
    session: Session,
    sheet_data: dict[str, list[dict[str, Any]]],
    course_id: int,
    create_by: int,
    tmpl: TemplateDef,
) -> ImportResult:
    """导入模板1数据：课程测试各题扣分情况 → CourseTestDetail。

    每位学生每条测试创建一条 CourseTestDetail 记录，
    包含 5 大题扣分 + 知识点 + 总成绩。
    同时自动识别并创建知识点。
    """
    result = ImportResult(detected_template=tmpl.name)

    # 收集所有行数据
    all_rows: list[dict[str, Any]] = []
    for sheet_name, rows in sheet_data.items():
        result.sheets_processed.append(sheet_name)
        all_rows.extend(rows)

    # 提取学期（优先取文件中的"学期"列，若为空则回退到课程默认学期）
    semester = _extract_semester(all_rows)
    if not semester:
        course = session.get(Course, course_id)
        semester = course.semester if course else ""

    # 第一遍：预先收集并创建所有知识点
    _collect_and_ensure_knowledge_points(session, course_id, all_rows)

    # 第二遍：导入数据行
    for sheet_name, rows in sheet_data.items():
        for row_data in rows:
            excel_row = row_data.get("_excel_row", 0)

            errors = validate_row(row_data, tmpl, sheet_name, excel_row)
            if errors:
                result.errors.extend(errors)
                result.error_count += len(errors)
                continue

            student_no = str(row_data.get("学号", "")).strip()
            student = _get_student_by_no(session, student_no)
            if not student:
                result.errors.append(ImportError(
                    sheet=sheet_name, row=excel_row, field="学号",
                    message=f"学号「{student_no}」在系统中不存在",
                ))
                result.error_count += 1
                continue

            _ensure_course_student(session, course_id, student.student_id)

            exam_name = str(row_data.get("测试名称") or "考试").strip()
            # 过滤内部字段和不需要的列
            source = {k: v for k, v in row_data.items()
                      if not k.startswith("_") and k != "所考查的知识点"}
            source_json = json.dumps(source, ensure_ascii=False, default=str)

            # 获取或创建考核批次
            batch_type = _map_batch_type_by_name(exam_name)
            default_weight = _get_default_batch_weight(exam_name, batch_type)
            batch = _get_or_create_exam_batch(
                session, course_id, exam_name,
                batch_type=batch_type, semester=semester, create_by=create_by,
                batch_weight=default_weight,
            )

            # 解析各题数据
            q_scores: list[float | None] = []
            q_knowledge: list[str | None] = []
            for qn in range(1, 6):
                col_score = f"第{qn}大题"
                col_knowledge = f"第 {qn} 大题扣分的主要知识点"
                q_scores.append(_safe_float(row_data.get(col_score)))
                kn = str(row_data.get(col_knowledge, "")).strip() if row_data.get(col_knowledge) else ""
                q_knowledge.append(kn if kn else None)

            total_score = _safe_float(row_data.get("总成绩")) or 0.0

            # Upsert: 查找是否已有该学生同批次的记录
            existing = session.exec(
                select(CourseTestDetail).where(
                    CourseTestDetail.student_id == student.student_id,
                    CourseTestDetail.exam_batch_id == batch.batch_id,
                )
            ).first()

            if existing:
                existing.question1_score = q_scores[0]
                existing.question2_score = q_scores[1]
                existing.question3_score = q_scores[2]
                existing.question4_score = q_scores[3]
                existing.question5_score = q_scores[4]
                existing.question1_knowledge = q_knowledge[0]
                existing.question2_knowledge = q_knowledge[1]
                existing.question3_knowledge = q_knowledge[2]
                existing.question4_knowledge = q_knowledge[3]
                existing.question5_knowledge = q_knowledge[4]
                existing.total_score = float(total_score)
                existing.source_data = source_json
                existing.update_time = datetime.now()
                session.add(existing)
            else:
                session.add(CourseTestDetail(
                    student_id=student.student_id,
                    exam_batch_id=batch.batch_id,
                    question1_score=q_scores[0],
                    question2_score=q_scores[1],
                    question3_score=q_scores[2],
                    question4_score=q_scores[3],
                    question5_score=q_scores[4],
                    question1_knowledge=q_knowledge[0],
                    question2_knowledge=q_knowledge[1],
                    question3_knowledge=q_knowledge[2],
                    question4_knowledge=q_knowledge[3],
                    question5_knowledge=q_knowledge[4],
                    total_score=float(total_score),
                    source_data=source_json,
                    create_by=create_by,
                ))
            result.success_count += 1

        # 每个 Sheet 处理完毕后统一提交，减少数据库事务次数
        session.commit()

    return result


# ============================================================================
# 模板2：单项成绩
# ============================================================================

def _import_simple_score(
    session: Session,
    sheet_data: dict[str, list[dict[str, Any]]],
    course_id: int,
    create_by: int,
    tmpl: TemplateDef,
) -> ImportResult:
    """导入单项成绩 → IndividualScore。

    格式：编号, 课程号, 课程名称, 学期, 学号, 姓名, 成绩名称, 成绩
    按"成绩名称"分组创建考试批次，每位学生每条成绩名称一条记录。
    """
    result = ImportResult(detected_template=tmpl.name)

    # 收集所有行
    all_rows: list[dict[str, Any]] = []
    for sheet_name, rows in sheet_data.items():
        result.sheets_processed.append(sheet_name)
        all_rows.extend(rows)

    # 优先取文件中的"学期"列，若为空则回退到课程默认学期
    semester = _extract_semester(all_rows)
    if not semester:
        course = session.get(Course, course_id)
        semester = course.semester if course else ""

    for sheet_name, rows in sheet_data.items():
        for row_data in rows:
            excel_row = row_data.get("_excel_row", 0)

            errors = validate_row(row_data, tmpl, sheet_name, excel_row)
            if errors:
                result.errors.extend(errors)
                result.error_count += len(errors)
                continue

            student_no = str(row_data.get("学号", "")).strip()
            student = _get_student_by_no(session, student_no)
            if not student:
                result.errors.append(ImportError(
                    sheet=sheet_name, row=excel_row, field="学号",
                    message=f"学号「{student_no}」在系统中不存在",
                ))
                result.error_count += 1
                continue

            _ensure_course_student(session, course_id, student.student_id)

            score_name = str(row_data.get("成绩名称", "")).strip()
            score_val = _safe_float(row_data.get("成绩"))
            if score_val is None:
                result.errors.append(ImportError(
                    sheet=sheet_name, row=excel_row, field="成绩",
                    message="成绩为空或格式错误",
                ))
                result.error_count += 1
                continue

            source = {k: v for k, v in row_data.items() if not k.startswith("_")}
            source_json = json.dumps(source, ensure_ascii=False, default=str)

            # 获取或创建考试批次
            batch_type = _map_batch_type_by_name(score_name)
            default_weight = _get_default_batch_weight(score_name, batch_type)
            batch = _get_or_create_exam_batch(
                session, course_id, score_name,
                batch_type=batch_type, semester=semester, create_by=create_by,
                batch_weight=default_weight,
            )

            # Upsert
            existing = session.exec(
                select(IndividualScore).where(
                    IndividualScore.student_id == student.student_id,
                    IndividualScore.exam_batch_id == batch.batch_id,
                )
            ).first()

            if existing:
                existing.score = float(score_val)
                existing.source_data = source_json
                existing.update_time = datetime.now()
                session.add(existing)
            else:
                session.add(IndividualScore(
                    student_id=student.student_id,
                    exam_batch_id=batch.batch_id,
                    score=float(score_val),
                    source_data=source_json,
                    create_by=create_by,
                ))
            result.success_count += 1

        # 每个 Sheet 处理完毕后统一提交，减少数据库事务次数
        session.commit()

    return result


# ============================================================================
# 模板3：成绩考勤情况
# ============================================================================

def _import_attendance(
    session: Session,
    sheet_data: dict[str, list[dict[str, Any]]],
    course_id: int,
    create_by: int,
    tmpl: TemplateDef,
) -> ImportResult:
    """导入成绩考勤情况 → AttendanceSheet。

    每位学生一行，包含 32 次考勤槽位 + 汇总统计。
    """
    result = ImportResult(detected_template=tmpl.name)

    # 收集所有行
    all_rows: list[dict[str, Any]] = []
    for sheet_name, rows in sheet_data.items():
        result.sheets_processed.append(sheet_name)
        all_rows.extend(rows)

    semester = _extract_semester(all_rows)

    # 考勤批次名称
    course = session.get(Course, course_id)
    if not semester:
        semester = course.semester if course else ""
    batch_name = f"考勤情况" if not course else f"{course.course_name}-考勤情况"

    batch = _get_or_create_exam_batch(
        session, course_id, batch_name,
        batch_type=5, semester=semester, create_by=create_by,
        batch_weight=5.0,  # 考勤默认权重
    )

    for sheet_name, rows in sheet_data.items():
        for row_data in rows:
            excel_row = row_data.get("_excel_row", 0)

            errors = validate_row(row_data, tmpl, sheet_name, excel_row)
            if errors:
                result.errors.extend(errors)
                result.error_count += len(errors)
                continue

            student_no = str(row_data.get("学号", "")).strip()
            student = _get_student_by_no(session, student_no)
            if not student:
                result.errors.append(ImportError(
                    sheet=sheet_name, row=excel_row, field="学号",
                    message=f"学号「{student_no}」在系统中不存在",
                ))
                result.error_count += 1
                continue

            _ensure_course_student(session, course_id, student.student_id)

            source = {k: v for k, v in row_data.items() if not k.startswith("_")}
            source_json = json.dumps(source, ensure_ascii=False, default=str)

            # 解析 32 次考勤
            att_values: list[str | None] = []
            for i in range(1, 33):
                col_name = f"考勤{i}"
                att_val = str(row_data.get(col_name, "")).strip() if row_data.get(col_name) else ""
                att_values.append(att_val if att_val else None)

            # Upsert
            existing = session.exec(
                select(AttendanceSheet).where(
                    AttendanceSheet.student_id == student.student_id,
                    AttendanceSheet.exam_batch_id == batch.batch_id,
                )
            ).first()

            if existing:
                _update_attendance_sheet(existing, att_values, row_data, source_json)
                session.add(existing)
            else:
                sheet = AttendanceSheet(
                    student_id=student.student_id,
                    exam_batch_id=batch.batch_id,
                    source_data=source_json,
                    create_by=create_by,
                )
                _update_attendance_sheet(sheet, att_values, row_data, source_json)
                session.add(sheet)

            result.success_count += 1

        # 每个 Sheet 处理完毕后统一提交，减少数据库事务次数
        session.commit()

    return result


def _update_attendance_sheet(
    sheet: AttendanceSheet,
    att_values: list[str | None],
    row_data: dict[str, Any],
    source_json: str,
) -> None:
    """将解析的考勤值填充到 AttendanceSheet 对象。"""
    # 32 次考勤
    sheet.attendance_1 = att_values[0]
    sheet.attendance_2 = att_values[1]
    sheet.attendance_3 = att_values[2]
    sheet.attendance_4 = att_values[3]
    sheet.attendance_5 = att_values[4]
    sheet.attendance_6 = att_values[5]
    sheet.attendance_7 = att_values[6]
    sheet.attendance_8 = att_values[7]
    sheet.attendance_9 = att_values[8]
    sheet.attendance_10 = att_values[9]
    sheet.attendance_11 = att_values[10]
    sheet.attendance_12 = att_values[11]
    sheet.attendance_13 = att_values[12]
    sheet.attendance_14 = att_values[13]
    sheet.attendance_15 = att_values[14]
    sheet.attendance_16 = att_values[15]
    sheet.attendance_17 = att_values[16]
    sheet.attendance_18 = att_values[17]
    sheet.attendance_19 = att_values[18]
    sheet.attendance_20 = att_values[19]
    sheet.attendance_21 = att_values[20]
    sheet.attendance_22 = att_values[21]
    sheet.attendance_23 = att_values[22]
    sheet.attendance_24 = att_values[23]
    sheet.attendance_25 = att_values[24]
    sheet.attendance_26 = att_values[25]
    sheet.attendance_27 = att_values[26]
    sheet.attendance_28 = att_values[27]
    sheet.attendance_29 = att_values[28]
    sheet.attendance_30 = att_values[29]
    sheet.attendance_31 = att_values[30]
    sheet.attendance_32 = att_values[31]

    # 汇总列
    sheet.total_count = _safe_int(row_data.get("考勤总数"))
    sheet.present_count = _safe_int(row_data.get("到课数"))
    sheet.leave_count = _safe_int(row_data.get("请假数"))
    sheet.late_count = _safe_int(row_data.get("迟到数"))
    sheet.early_leave_count = _safe_int(row_data.get("早退数"))
    sheet.attendance_rate = _safe_float(row_data.get("到课率"))

    sheet.source_data = source_json
    sheet.update_time = datetime.now()


# ============================================================================
# 导入调度
# ============================================================================

IMPORT_HANDLERS: dict[str, Callable] = {
    TEMPLATE_EXAM_DEDUCTION.template_id: _import_exam_deduction,
    TEMPLATE_ATTENDANCE.template_id: _import_attendance,
    TEMPLATE_SIMPLE_SCORE.template_id: _import_simple_score,
}


def import_file(
    session: Session,
    file_path: str,
    file_ext: str,
    course_id: int,
    create_by: int,
) -> ImportResult:
    """主导入流程：解析文件 → 按 Sheet 检测模板 → 校验 → 导入。

    每个 Sheet 独立检测模板并导入，同一文件可包含多种模板数据。
    """
    # 1. 解析文件
    if file_ext == ".xlsx":
        sheet_data = parse_xlsx(file_path)
    elif file_ext == ".txt":
        sheet_data = parse_txt(file_path)
    else:
        result = ImportResult()
        result.errors.append(ImportError(
            sheet="", row=0, field=None,
            message=f"不支持的文件格式「{file_ext}」，仅支持 .xlsx 和 .txt",
        ))
        result.error_count = 1
        return result

    if not sheet_data:
        result = ImportResult()
        result.errors.append(ImportError(
            sheet="", row=0, field=None,
            message="文件为空或无法解析",
        ))
        result.error_count = 1
        return result

    # 2. 按 Sheet 逐表检测模板并分组
    sheet_templates: dict[str, tuple[TemplateDef, list[dict[str, Any]]]] = {}
    unmatchable_sheets: list[str] = []

    for sheet_name, rows in sheet_data.items():
        if not rows:
            continue

        headers = [k for k in rows[0].keys() if not k.startswith("_")]
        tmpl = detect_template(headers)

        if tmpl:
            sheet_templates[sheet_name] = (tmpl, rows)
        else:
            unmatchable_sheets.append(sheet_name)

    if not sheet_templates:
        all_headers = {}
        for sname in unmatchable_sheets:
            if sheet_data[sname]:
                all_headers[sname] = [k for k in sheet_data[sname][0].keys() if not k.startswith("_")]

        result = ImportResult()
        result.errors.append(ImportError(
            sheet="", row=0, field=None,
            message=(
                f"无法识别文件模板。支持的模板：课程测试各题扣分情况、单项成绩、成绩考勤情况。"
                f"各Sheet表头：{all_headers}"
            ),
        ))
        result.error_count = 1
        return result

    # 3. 合并结果（按模板分组，同一模板的 sheet 批量导入）
    merged = ImportResult()
    merged.detected_template = ", ".join(
        sorted({tmpl.name for tmpl, _ in sheet_templates.values()})
    )

    # 按 template_id 分组
    by_template: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for sname, (tmpl, rows) in sheet_templates.items():
        by_template.setdefault(tmpl.template_id, {})[sname] = rows

    for template_id, sheets in by_template.items():
        handler = IMPORT_HANDLERS.get(template_id)
        if not handler:
            merged.errors.append(ImportError(
                sheet="", row=0, field=None,
                message=f"模板 ID「{template_id}」的导入逻辑尚未实现",
            ))
            merged.error_count += 1
            continue

        # 取第一个 sheet 对应的模板定义
        first_tmpl = sheet_templates[next(iter(sheets))][0]
        partial = handler(session, sheets, course_id, create_by, first_tmpl)
        merged.success_count += partial.success_count
        merged.error_count += partial.error_count
        merged.errors.extend(partial.errors)
        merged.sheets_processed.extend(partial.sheets_processed)

    # 记录未匹配的 sheet（警告级别，不影响匹配成功的 sheet）
    if unmatchable_sheets:
        for sname in unmatchable_sheets:
            merged.errors.append(ImportError(
                sheet=sname, row=0, field=None,
                message=f"Sheet「{sname}」的数据格式未能匹配任何已知模板，已跳过",
            ))
            merged.error_count += 1

    return merged


# ============================================================================
# 模板下载：生成标准模板文件
# ============================================================================

from openpyxl.styles import Font as _Font, Alignment as _Alignment, PatternFill as _PatternFill
from openpyxl.utils import get_column_letter as _get_column_letter

# 模板元信息（template_id、名称、分类、表头、示例行）
TEMPLATE_META: list[dict[str, Any]] = [
    {
        "template_id": TEMPLATE_EXAM_DEDUCTION.template_id,
        "name": "课程测试各题扣分情况",
        "dataType": "成绩",
        "description": "导入课程测试各题扣分明细，包含每位学生各大题扣分及对应知识点，适用于期中/期末考试扣分分析。",
        "headers": [
            "编号", "课程号", "课程名称", "学期", "测试名称", "学号", "姓名",
            "第1大题", "第 1 大题扣分的主要知识点",
            "第2大题", "第 2 大题扣分的主要知识点",
            "第3大题", "第 3 大题扣分的主要知识点",
            "第4大题", "第 4 大题扣分的主要知识点",
            "第5大题", "第 5 大题扣分的主要知识点",
            "总成绩",
        ],
        "example": [
            1, "CS101", "计算机网络", "2025-2026-1", "期中考试", "2024001001", "赵伟",
            2, "数据报分片", 2, "TCP/IP参考模型", 2, "TCP报文段",
            2, "可靠传输技术", 2, "ARP协议", 90,
        ],
        "instruction": "填写说明：①总成绩为实际得分（满分100）；②第1~5大题为扣分值（非得分）；③知识点填写该题扣分对应的知识领域名称。",
    },
    {
        "template_id": TEMPLATE_SIMPLE_SCORE.template_id,
        "name": "单项成绩",
        "dataType": "成绩",
        "description": "导入课程单项成绩，每行一条成绩记录，按成绩名称区分不同考试/考查项目。",
        "headers": [
            "编号", "课程号", "课程名称", "学期", "学号", "姓名",
            "成绩名称", "成绩",
        ],
        "example": [
            1, "CS101", "计算机网络", "2025-2026-1", "2024001001", "赵伟",
            "期中考试", 85,
        ],
        "instruction": "填写说明：①成绩名称填入考试/考查项目名（如期中考试、期末考试、作业等），同一课程同名学生可有多条记录；②成绩为百分制 0-100 的数字。",
    },
    {
        "template_id": TEMPLATE_ATTENDANCE.template_id,
        "name": "成绩考勤情况",
        "dataType": "考勤",
        "description": "导入课程考勤记录，每位学生32次课的出勤状态（到/缺/请假/迟到/早退）。",
        "headers": [
            "编号", "课程号", "课程名称", "学期", "学号", "姓名",
            *[f"考勤{i}" for i in range(1, 33)],
            "考勤总数", "到课数", "请假数", "早退数", "迟到数", "到课率",
        ],
        "example": [
            1, "CS101", "计算机网络", "2025-2026-1", "2024001001", "赵伟",
            *(["到"] * 32),
            None, None, None, None, None, None,
        ],
        "instruction": "填写说明：①考勤列填「到」「缺」「请假」「迟到」「早退」之一；②考勤总数、到课数、请假数、早退数、迟到数、到课率为Excel公式自动计算，请勿手动填写。",
    },
]


def generate_template_xlsx(template_id: str) -> bytes:
    """根据模板 ID 生成标准 .xlsx 模板文件（仅含表头+示例行+填写说明）。"""
    meta = next((m for m in TEMPLATE_META if m["template_id"] == template_id), None)
    if not meta:
        raise ValueError(f"未知模板 ID: {template_id}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 样式
    header_font = _Font(bold=True, size=11)
    header_fill = _PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    example_fill = _PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    wrap_align = _Alignment(wrap_text=True, vertical="center")

    headers = meta["headers"]
    example = meta["example"]

    # Row 1: 表头
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_align

    # Row 2: 示例数据（黄色底色标示）
    for ci, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=ci, value=val)
        cell.fill = example_fill

    # Row 4: 填写说明
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=len(headers))
    instr_cell = ws.cell(row=4, column=1, value=meta["instruction"])
    instr_cell.font = _Font(italic=True, size=10, color="666666")

    # 调整列宽
    for ci, h in enumerate(headers, start=1):
        col_letter = _get_column_letter(ci)
        width = max(len(str(h)) * 1.5, 10)
        if width > 30:
            width = 30
        ws.column_dimensions[col_letter].width = width

    output = _io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_template_txt(template_id: str) -> bytes:
    """根据模板 ID 生成 UTF-8 逗号分隔 .txt 模板文件（含表头+示例行+填写说明）。"""
    meta = next((m for m in TEMPLATE_META if m["template_id"] == template_id), None)
    if not meta:
        raise ValueError(f"未知模板 ID: {template_id}")

    headers = meta["headers"]
    example = meta["example"]
    instruction = meta["instruction"]

    lines: list[str] = []

    # 填写说明（以 # 开头作为注释行）
    lines.append(f"# {instruction}")
    lines.append(f"# 模板：{meta['name']}")

    # 表头行（逗号分隔）
    lines.append(", ".join(str(h) for h in headers))

    # 示例数据行（逗号分隔，空值用空字符串）
    lines.append(", ".join(str(v) if v is not None else "" for v in example))

    content = "\n".join(lines)

    # UTF-8 BOM + 内容（确保 Excel 能正确识别中文）
    return b"\xef\xbb\xbf" + content.encode("utf-8")
