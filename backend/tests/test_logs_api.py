"""操作日志 API 测试:筛选 / 分页 / 导出 / 权限边界。

使用独立的内存 SQLite 引擎,避免污染 conftest 的共享测试库。
"""
import io as _io
from datetime import datetime, timedelta

import openpyxl
import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from sqlalchemy.pool import StaticPool
from sqlmodel import Session, SQLModel, create_engine

from app import models  # noqa: F401  确保所有表注册到 metadata
from app.api.v1.auth import create_token
from app.api.v1.logs import router as logs_router
from app.core.database import get_session
from app.models import SysOperationLog, SysRole, SysUser

NOW = datetime(2026, 8, 31, 10, 0, 0)


@pytest.fixture(scope="module")
def engine():
    # StaticPool:所有 Session 共享同一连接,保证 :memory: 库在 TestClient
    # 的工作线程中也能看到同一份数据
    eng = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    SQLModel.metadata.create_all(eng)
    with Session(eng) as s:
        s.add(SysRole(role_id=1, role_name="系统管理员", role_code="admin"))
        s.add(SysRole(role_id=2, role_name="教师", role_code="teacher"))
        s.add(SysUser(user_id=1, username="admin", password="x", real_name="管理员", role_id=1, status=1))
        s.add(SysUser(user_id=2, username="teacher1", password="x", real_name="王老师", role_id=2, status=1))
        s.add(SysUser(user_id=3, username="teacher2", password="x", real_name="李老师", role_id=2, status=1))
        logs = [
            SysOperationLog(user_id=1, module="用户管理", operation="新增",
                            content="创建用户:u1", ip_address="127.0.0.1",
                            operation_time=NOW),
            SysOperationLog(user_id=2, module="课程管理", operation="新增",
                            content="新增课程:CS1", ip_address="127.0.0.1",
                            operation_time=NOW - timedelta(days=1)),
            SysOperationLog(user_id=2, module="数据管理", operation="导入",
                            content="导入成绩", ip_address="127.0.0.1",
                            operation_time=NOW - timedelta(days=2)),
            SysOperationLog(user_id=3, module="课程管理", operation="编辑",
                            content="更新课程:CS1", ip_address="127.0.0.1",
                            operation_time=NOW - timedelta(days=3)),
            SysOperationLog(user_id=1, module="用户管理", operation="删除",
                            content="删除用户:u2", ip_address="127.0.0.1",
                            operation_time=NOW - timedelta(days=4)),
        ]
        s.add_all(logs)
        s.commit()
    return eng


@pytest.fixture
def client(engine):
    app = FastAPI()
    app.include_router(logs_router, prefix="/api/v1")

    def override_session():
        with Session(engine) as s:
            yield s

    app.dependency_overrides[get_session] = override_session
    return TestClient(app)


def _auth(user_id: int, username: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {create_token(user_id, username)}"}


ADMIN = _auth(1, "admin")


def test_list_requires_admin(client):
    """非管理员访问日志列表返回 403。"""
    resp = client.get("/api/v1/logs", headers=_auth(2, "teacher1"))
    assert resp.status_code == 403


def test_pagination(client):
    """分页:total 正确,每页条数与翻页边界正确。"""
    resp = client.get("/api/v1/logs", params={"page": 1, "page_size": 2}, headers=ADMIN)
    assert resp.status_code == 200
    body = resp.json()
    assert body["total"] == 5
    assert len(body["list"]) == 2

    # 按时间倒序,第一页第一条是最新日志
    assert body["list"][0]["operation"] == "创建用户:u1"

    resp3 = client.get("/api/v1/logs", params={"page": 3, "page_size": 2}, headers=ADMIN)
    assert len(resp3.json()["list"]) == 1

    # 超出范围的页码返回空列表,total 不变
    resp9 = client.get("/api/v1/logs", params={"page": 9, "page_size": 2}, headers=ADMIN)
    assert resp9.json()["list"] == []
    assert resp9.json()["total"] == 5


def test_filter_by_username(client):
    """按用户名模糊筛选。"""
    resp = client.get("/api/v1/logs", params={"username": "teacher1"}, headers=ADMIN)
    body = resp.json()
    assert body["total"] == 2
    assert all(item["username"] == "teacher1" for item in body["list"])

    # 模糊匹配子串(teacher1 两条 + teacher2 一条)
    resp2 = client.get("/api/v1/logs", params={"username": "teacher"}, headers=ADMIN)
    assert resp2.json()["total"] == 3

    # 无匹配用户返回空
    resp3 = client.get("/api/v1/logs", params={"username": "不存在的人"}, headers=ADMIN)
    assert resp3.json()["total"] == 0


def test_filter_by_module_and_operation(client):
    """按操作模块 / 操作类型精确筛选,可组合。"""
    resp = client.get("/api/v1/logs", params={"module": "课程管理"}, headers=ADMIN)
    assert resp.json()["total"] == 2

    resp = client.get("/api/v1/logs", params={"operation": "新增"}, headers=ADMIN)
    assert resp.json()["total"] == 2

    resp = client.get(
        "/api/v1/logs",
        params={"module": "用户管理", "operation": "新增", "username": "admin"},
        headers=ADMIN,
    )
    body = resp.json()
    assert body["total"] == 1
    assert body["list"][0]["operationType"] == "新增"


def test_filter_by_date_range(client):
    """按时间范围筛选(含边界)。"""
    resp = client.get(
        "/api/v1/logs",
        params={"start_date": "2026-08-29", "end_date": "2026-08-31"},
        headers=ADMIN,
    )
    assert resp.json()["total"] == 3

    # 仅开始日期
    resp = client.get("/api/v1/logs", params={"start_date": "2026-08-31"}, headers=ADMIN)
    assert resp.json()["total"] == 1

    # 仅结束日期
    resp = client.get("/api/v1/logs", params={"end_date": "2026-08-27"}, headers=ADMIN)
    assert resp.json()["total"] == 1

    # 空区间
    resp = client.get(
        "/api/v1/logs",
        params={"start_date": "2026-08-01", "end_date": "2026-08-02"},
        headers=ADMIN,
    )
    assert resp.json()["total"] == 0


def test_list_modules(client):
    """模块列表接口返回去重后的模块名。"""
    resp = client.get("/api/v1/logs/modules", headers=ADMIN)
    assert resp.status_code == 200
    assert set(resp.json()["list"]) == {"用户管理", "课程管理", "数据管理"}


def test_export_all(client):
    """导出全部日志为 Excel,内容与列表一致。"""
    resp = client.get("/api/v1/logs/export", headers=ADMIN)
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "operation_logs_export.xlsx" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # 表头 + 5 条数据
    assert len(rows) == 6
    assert rows[0] == ("序号", "操作用户", "操作模块", "操作类型", "操作内容", "IP 地址", "操作时间")
    # 按时间倒序,第一条数据为最新日志
    assert rows[1][1] == "admin"
    assert rows[1][4] == "创建用户:u1"


def test_export_with_filters(client):
    """导出遵循筛选条件。"""
    resp = client.get(
        "/api/v1/logs/export",
        params={"module": "课程管理", "start_date": "2026-08-29"},
        headers=ADMIN,
    )
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    rows = list(wb.active.iter_rows(values_only=True))
    # 课程管理 2 条中,8-29 及之后仅 1 条(8-30)
    assert len(rows) == 2
    assert rows[1][2] == "课程管理"
